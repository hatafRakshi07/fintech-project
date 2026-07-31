import openpyxl
import sys
from datetime import datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

EXCEL_PATH = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder (1).xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

gift_sheets = [
    "Shree krishna aasociates gift r",
    "Radhe krishna bissi gift list",
    "gift stock maintain"
]

for sname in gift_sheets:
    if sname in wb.sheetnames:
        print(f"\n=================== SHEET: {sname} ===================")
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        print(f"Total rows: {len(rows)}")
        if rows:
            print("Row 1 (Header):", [str(c) if c is not None else "" for c in rows[0][:15]])
            for idx in range(1, min(10, len(rows))):
                print(f"Row {idx+1}:", [str(c) if c is not None else "" for c in rows[idx][:15]])

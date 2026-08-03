import openpyxl
import psycopg2

wb = openpyxl.load_workbook('C:/Users/iSN_kota_T52/Downloads/Bissi folder (1).xlsx', data_only=True)
conn = psycopg2.connect('postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require')
cur = conn.cursor()

SCHEMES = [
    ('Hare ka sahara bissi 20 date',    'BISSI-1', '11111111-1111-1111-1111-111111111111', 7),
    ('Shree Krishna associate lottery', 'BISSI-2', '22222222-2222-2222-2222-222222222222', 6),
    ('Pyare mohan 15 date',             'BISSI-3', '33333333-3333-3333-3333-333333333333', 9),
    ('Sawariya seth 5 date',            'BISSI-4', 'a3d68b9c-63df-4884-a5ad-eb8a17e3be31', 7),
]

print('=== JSK FINAL STATUS REPORT ===\n')

for sheet_name, code, uuid, data_col in SCHEMES:
    ws = wb[sheet_name]
    header = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    date_cols = [i for i, v in enumerate(header) if v and hasattr(v,'year')]

    cur.execute('SELECT normalized_token_number FROM tokens WHERE committee_id=%s', [uuid])
    db_tokens = {row[0] for row in cur.fetchall()}

    has_payment = []
    no_payment  = []

    for r in range(2, ws.max_row+1):
        tok = ws.cell(r,1).value
        if not tok:
            continue
        try:
            t = int(float(str(tok)))
        except:
            continue
        if t in db_tokens:
            continue

        name = str(ws.cell(r,2).value or '').strip()
        ref  = str(ws.cell(r,3).value or '').strip()
        mob  = str(ws.cell(r,4).value or '').strip()

        pays = 0
        amt  = 0.0
        for ci in date_cols:
            v = ws.cell(r, ci+1).value
            if v and isinstance(v, (int, float)) and v > 0:
                pays += 1
                amt  += float(v)

        if pays > 0:
            has_payment.append((t, name, ref, mob, pays, amt))
        else:
            no_payment.append(t)

    print(f'{code}:')
    print(f'  Tokens with payments (MISSING from DB): {len(has_payment)}')
    print(f'  Tokens with NO payments (empty slots) : {len(no_payment)}')
    if has_payment:
        print('  --- Tokens with payments missing from DB ---')
        for t, nm, ref, mob, pays, amt in sorted(has_payment):
            n = nm if nm else '(koi naam nahi)'
            r2 = ref if ref else '(koi reference nahi)'
            m = mob if mob else '(mobile nahi)'
            print(f'    Token #{t}  naam={n}  ref={r2}  mobile={m}  {pays} payments  Rs{int(amt):,}')
    else:
        print('  Sab empty slots hain - koi payment nahi')
    print()

wb.close()
conn.close()

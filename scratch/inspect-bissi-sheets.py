import openpyxl

file_path = r"C:\Users\lenovo\Downloads\Bissi folder (5).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

print("Workbook sheet names:", wb.sheetnames)

bissi_sheets = [
    'Sawariya seth 5 date',
    'Pyare mohan 15 date',
    'Hare ka sahara bissi 20 date',
    'Shree Krishna associate lottery'
]

for sheet_name in wb.sheetnames:
    if any(b.lower() in sheet_name.lower() for b in bissi_sheets):
        ws = wb[sheet_name]
        print(f"\n========================================")
        print(f"SHEET: '{sheet_name}' (max_row: {ws.max_row}, max_col: {ws.max_column})")
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, min(25, ws.max_column + 1))]
            if any(v is not None for v in row_vals):
                print(f"Row {r}: {row_vals}")

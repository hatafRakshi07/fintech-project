import openpyxl
import json
import sys

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder (1).xlsx"
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
print("Sheets found:", wb.sheetnames)

summary = {}

for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in rows[0]] if rows else []
    
    print(f"\n--- Sheet: {sheet} ---")
    print(f"Total rows: {len(rows)}")
    print(f"Headers (Row 1): {headers[:15]}")
    if len(rows) > 1:
        print(f"Sample Row 2: {[str(cell) if cell is not None else '' for cell in rows[1][:15]]}")

    summary[sheet] = {
        "total_rows": len(rows),
        "headers": headers[:30],
        "sample_row": [str(c) if c is not None else "" for c in rows[1][:30]] if len(rows) > 1 else []
    }

with open(r"scratch\excel_summary.json", "w", encoding="utf-8") as f:
    json.dump(summary, f, indent=2, ensure_ascii=False)

print("\nSummary saved to scratch/excel_summary.json")

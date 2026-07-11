"""
Fast bulk import: Bissi folder.xlsx → Neon PostgreSQL
Direct DB insert (bypasses API for speed)
Run: python import-bissi-fast.py
"""
import re, datetime
import psycopg2
import psycopg2.extras
import openpyxl

DB_URL    = "postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed-pooler.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require"
XLSX_FILE = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder.xlsx"

def clean_mobile(val):
    if not val: return None
    s = re.sub(r"[^\d]", "", str(val).split(".")[0])
    s = s[-10:] if len(s) >= 10 else s
    return s if len(s) >= 6 else None

def clean_name(val):
    if not val: return None
    v = str(val).strip()
    return v[:200] if v and v.lower() not in ("none","name","name ","") else None

print("Connecting to Neon...")
conn = psycopg2.connect(DB_URL)
cur  = conn.cursor()
print("  Connected!\n")

# ── Branch ─────────────────────────────────────────────────────────────────
cur.execute("SELECT id FROM branches WHERE code = 'SKA001'")
row = cur.fetchone()
if row:
    BRANCH_ID = row[0]
    print(f"Using existing branch ID: {BRANCH_ID}")
else:
    cur.execute("""
        INSERT INTO branches (name, code, city, status, created_at, updated_at)
        VALUES ('Shree Krishna Associate', 'SKA001', 'Jaipur', 'active', NOW(), NOW())
        RETURNING id
    """)
    BRANCH_ID = cur.fetchone()[0]
    conn.commit()
    print(f"Created branch ID: {BRANCH_ID}")

# ── Read Excel ─────────────────────────────────────────────────────────────
print("\nReading Excel...")
wb = openpyxl.load_workbook(XLSX_FILE)

SHEETS = [
    ("Sawariya seth 5 date",           "Sawariya Seth Bissi",        5000, 0, 1, 3, 5, 2),
    ("Pyare mohan 15 date",            "Pyare Mohan Bissi",          5000, 0, 1, 3, 6, 2),
    ("Hare ka sahara bissi 20 date",   "Hare Ka Sahara Bissi",       2500, 0, 1, 3, 4, 2),
    ("Shree Krishna associate lottery","Shree Krishna Bissi",         3000, 0, 1, 3, 5, 2),
]

# Collect unique customers
customers = {}  # key -> dict
sheet_tokens = {}

for sheet_name, comm_name, amount, col_tok, col_name, col_mob, col_addr, col_ref in SHEETS:
    ws = wb[sheet_name]
    tokens = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name   = clean_name(row[col_name] if len(row)>col_name else None)
        mobile = clean_mobile(row[col_mob] if len(row)>col_mob else None)
        addr   = str(row[col_addr]).strip()[:500] if len(row)>col_addr and row[col_addr] else None
        ref    = clean_name(row[col_ref] if len(row)>col_ref else None)
        try:
            tok_raw = row[col_tok] if len(row)>col_tok and row[col_tok] else 0
            tok = int(float(str(tok_raw).split("(")[0].strip())) if tok_raw else 0
        except: tok = 0

        if not name: continue
        key = mobile if mobile else f"name:{name.lower()}"
        if key not in customers:
            customers[key] = (name, mobile, addr, ref)
        tokens.append((tok, key))
    sheet_tokens[sheet_name] = tokens
    print(f"  {sheet_name}: {len(tokens)} rows")

print(f"\n  Unique customers: {len(customers)}")

# ── Bulk insert customers ──────────────────────────────────────────────────
print("\nInserting customers (skipping already imported)...")
cur.execute("SELECT COUNT(*) FROM customers WHERE branch_id = %s", (BRANCH_ID,))
existing_count = cur.fetchone()[0]
if existing_count >= len(customers):
    print(f"  {existing_count} customers already in DB, skipping")
    cust_data = []
else:
    cust_data = []
    for i, (key, (name, mobile, addr, ref)) in enumerate(customers.items()):
        ref_num = f"SKA{str(existing_count + i + 1).zfill(5)}"
        cust_data.append((ref_num, name, mobile, addr, ref, BRANCH_ID))

    psycopg2.extras.execute_values(cur, """
        INSERT INTO customers (reference_number, name, mobile, address, reference_name, branch_id, status, created_at, updated_at)
        VALUES %s
        ON CONFLICT (reference_number) DO NOTHING
    """, [(r,n,m,a,rf,b,"active",datetime.datetime.now(),datetime.datetime.now()) for r,n,m,a,rf,b in cust_data],
    page_size=200)
    conn.commit()
    print(f"  Inserted {len(cust_data)} customers")

# Build mobile → DB id map
cur.execute("SELECT id, mobile FROM customers WHERE branch_id = %s", (BRANCH_ID,))
db_customers = cur.fetchall()
mobile_to_id = {row[1]: row[0] for row in db_customers if row[1]}

# ── Create Committees ──────────────────────────────────────────────────────
print("\nCreating committees...")
today = datetime.date.today()
comm_id_map = {}  # sheet_name → committee_id
for sheet_name, comm_name, amount, *_ in SHEETS:
    cur.execute("SELECT id FROM committees WHERE name = %s AND branch_id = %s", (comm_name, BRANCH_ID))
    existing = cur.fetchone()
    if existing:
        comm_id = existing[0]
        print(f"  Exists: {comm_name} (ID: {comm_id})")
    else:
        cur.execute("""
            INSERT INTO committees (name, type, installment_amount, member_limit, draw_date, duration, branch_id, status, created_at, updated_at)
            VALUES (%s, 'monthly', %s, 100, %s, 24, %s, 'active', NOW(), NOW())
            RETURNING id
        """, (comm_name, str(amount), today, BRANCH_ID))
        comm_id = cur.fetchone()[0]
        conn.commit()
        print(f"  Created: {comm_name} (ID: {comm_id})")
    comm_id_map[sheet_name] = comm_id

# ── Import Tokens (one per customer-per-token slot) ────────────────────────
print("\nImporting tokens...")
total_tokens = 0

for sheet_name, comm_name, amount, col_tok, col_name, col_mob, col_addr, col_ref in SHEETS:
    comm_id = comm_id_map.get(sheet_name)
    if not comm_id: continue

    # Check existing tokens for this committee
    cur.execute("SELECT COUNT(*) FROM tokens WHERE committee_id = %s", (comm_id,))
    existing_tokens = cur.fetchone()[0]
    if existing_tokens > 0:
        print(f"  {comm_name}: {existing_tokens} tokens already imported, skipping")
        continue

    ws = wb[sheet_name]
    token_data = []
    name_to_cust_id = {}

    # Build name → customer_id map for this sheet (fallback for missing mobile)
    cur.execute("""
        SELECT id, name, mobile FROM customers WHERE branch_id = %s
    """, (BRANCH_ID,))
    for cid, cname, cmob in cur.fetchall():
        name_to_cust_id[cname.lower().strip()] = cid

    for row in ws.iter_rows(min_row=2, values_only=True):
        name   = clean_name(row[col_name] if len(row)>col_name else None)
        mobile = clean_mobile(row[col_mob] if len(row)>col_mob else None)
        try:
            tok_raw = row[col_tok] if len(row)>col_tok and row[col_tok] else None
            tok = str(int(float(str(tok_raw).split("(")[0].strip()))) if tok_raw else None
        except: tok = None

        if not name or not tok: continue

        # Resolve customer id
        cust_id = mobile_to_id.get(mobile) if mobile else None
        if not cust_id and name:
            cust_id = name_to_cust_id.get(name.lower().strip())
        if not cust_id: continue

        token_data.append((tok, cust_id, comm_id))

    if token_data:
        psycopg2.extras.execute_values(cur, """
            INSERT INTO tokens (token_number, customer_id, committee_id, status, created_at, updated_at)
            VALUES %s
            ON CONFLICT DO NOTHING
        """, [(t, c, cm, "active", datetime.datetime.now(), datetime.datetime.now())
              for t, c, cm in token_data], page_size=200)
        conn.commit()
        total_tokens += len(token_data)
        print(f"  {comm_name}: {len(token_data)} tokens imported")

print(f"\n  Total tokens imported: {total_tokens}")

conn.commit()
cur.close()
conn.close()

print("\n" + "="*50)
print("IMPORT COMPLETE!")
print(f"  Branch:     Shree Krishna Associate (ID: {BRANCH_ID})")
print(f"  Customers:  {len(cust_data) or existing_count} total")
print(f"  Committees: {len(comm_id_map)} created")
print(f"  Tokens:     {total_tokens} imported")
print("="*50)
print("\nOpen the app and check Customers + Committees + Tokens tabs!")

# -*- coding: utf-8 -*-
"""
Extract ALL real data from Bissi folder.xlsx into JSON files.
Outputs:
  - extracted_customers.json   (de-duped across all sheets)
  - extracted_committees.json  (4 main bissi committees)
  - extracted_tokens.json      (token + committee member mapping)
  - extracted_collections.json (payments from date columns)
  - extracted_interests.json   (BYAJ KI LIST)
  - extracted_loans.json       (nikku ji loan)
  - extracted_gifts.json       (gift distribution records)
  - extracted_lotteries.json   (lucky draw data)
  - extracted_daily_collections.json (Daily collection + office 2)
"""

import openpyxl
import json
import re
import sys
import os
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder.xlsx"
OUTPUT_DIR = r"C:\Users\iSN_kota_T52\Desktop\File-Processor"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ─── Helpers ──────────────────────────────────────────────────────────────────

def clean_str(v):
    """Clean cell value to string or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s.lower() in ('none', '', 'nan', 'null'):
        return None
    return s

def clean_mobile(v):
    """Clean mobile number."""
    s = clean_str(v)
    if not s:
        return None
    # Remove .0 suffix from float
    s = re.sub(r'\.0$', '', s)
    # Remove non-digit chars
    digits = re.sub(r'\D', '', s)
    if len(digits) >= 10:
        return digits[-10:]  # last 10 digits
    return s if len(s) > 3 else None

def clean_amount(v):
    """Parse amount from cell (could be number or string)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip()
    # Remove commas
    s = s.replace(',', '')
    # Try to extract number
    m = re.search(r'[\d.]+', s)
    if m:
        try:
            return round(float(m.group()), 2)
        except:
            return None
    return None

def date_to_str(v):
    """Convert date cell to YYYY-MM-DD string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    # Try common formats
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%y', '%d/%m/%Y']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except:
            pass
    return None

def make_customer_key(name, mobile):
    """Create a de-duplication key for customers."""
    n = (name or '').strip().lower()
    m = clean_mobile(mobile) or ''
    return f"{n}|{m}"

# ─── 1. Extract Committee Members (Customers + Tokens) ──────────────────────

print("=== Phase 1: Extracting committee members ===")

COMMITTEE_SHEETS = [
    {
        'sheet': 'Sawariya seth 5 date',
        'name': 'Sawariya Seth Bissi',
        'draw_day': 5,
        'installment': 3000,
        'token_col': 1,   # TOKEN
        'name_col': 2,    # Name
        'ref_col': 3,     # Reference name
        'mobile_col': 4,  # Mobile No
        'ref_mobile_col': 5,  # Reference mobile
        'address_col': 6, # Address
        'reason_col': 7,  # Reason
        'data_start_col': 8,  # Date columns start
    },
    {
        'sheet': 'Pyare mohan 15 date',
        'name': 'Pyare Mohan Bissi',
        'draw_day': 15,
        'installment': 3000,
        'token_col': 1,
        'name_col': 2,
        'ref_col': 3,
        'mobile_col': 4,
        'ref_mobile_col': 5,
        'address_col': 7,  # Address is col 7
        'reason_col': 6,   # Reason is col 6
        'reply_col': 8,    # Reply col
        'data_start_col': 9,
    },
    {
        'sheet': 'Hare ka sahara bissi 20 date',
        'name': 'Hare Ka Sahara Bissi',
        'draw_day': 20,
        'installment': 2500,
        'token_col': 1,
        'name_col': 2,
        'ref_col': 3,
        'mobile_col': 4,
        'address_col': 5,
        'reason_col': 7,
        'data_start_col': 8,
    },
    {
        'sheet': 'Shree Krishna associate lottery',
        'name': 'Shree Krishna Associates',
        'draw_day': 25,
        'installment': 3000,
        'token_col': 1,
        'name_col': 2,
        'ref_col': 3,
        'mobile_col': 4,
        'ref_mobile_col': 5,
        'address_col': 6,
        'reason_col': 7,
        'data_start_col': 8,
    },
]

all_customers = {}  # key -> customer dict
all_tokens = []     # committee_idx, token_number, customer_key
all_collections = []  # payment records from date columns
committees_data = []

for ci, cinfo in enumerate(COMMITTEE_SHEETS):
    sname = cinfo['sheet']
    print(f"\n  Processing: {sname}")
    ws = wb[sname]
    
    committees_data.append({
        'idx': ci,
        'name': cinfo['name'],
        'installment': cinfo['installment'],
        'draw_day': cinfo['draw_day'],
        'type': 'monthly',
    })
    
    # Find date columns in header row
    date_columns = []
    for c in range(cinfo['data_start_col'], ws.max_column + 1):
        hdr = ws.cell(1, c).value
        d = date_to_str(hdr)
        if d:
            date_columns.append((c, d))
    
    print(f"    Found {len(date_columns)} date columns")
    
    member_count = 0
    for r in range(2, ws.max_row + 1):
        token_val = ws.cell(r, cinfo['token_col']).value
        name_val = clean_str(ws.cell(r, cinfo['name_col']).value)
        
        # Skip empty rows
        if token_val is None and name_val is None:
            continue
        
        token_num = clean_str(token_val)
        if token_num:
            token_num = re.sub(r'\.0$', '', token_num)
        
        if not name_val and not token_num:
            continue
            
        mobile = clean_mobile(ws.cell(r, cinfo['mobile_col']).value)
        ref_name = clean_str(ws.cell(r, cinfo['ref_col']).value)
        ref_mobile = clean_mobile(ws.cell(r, cinfo.get('ref_mobile_col', 0)).value) if cinfo.get('ref_mobile_col') else None
        address = clean_str(ws.cell(r, cinfo.get('address_col', 0)).value) if cinfo.get('address_col') else None
        reason = clean_str(ws.cell(r, cinfo.get('reason_col', 0)).value) if cinfo.get('reason_col') else None
        
        # Customer key for de-duplication
        ckey = make_customer_key(name_val or f"Token-{token_num}", mobile)
        
        if ckey not in all_customers:
            all_customers[ckey] = {
                'name': name_val or f"Token-{token_num}",
                'mobile': mobile or 'N/A',
                'reference_name': ref_name,
                'reference_mobile': ref_mobile,
                'address': address,
                'reason': reason,
                'committees': [],
            }
        else:
            # Update with more complete info if available
            existing = all_customers[ckey]
            if not existing.get('address') and address:
                existing['address'] = address
            if not existing.get('reference_name') and ref_name:
                existing['reference_name'] = ref_name
            if not existing.get('reference_mobile') and ref_mobile:
                existing['reference_mobile'] = ref_mobile
        
        all_customers[ckey]['committees'].append(cinfo['name'])
        
        # Token mapping
        all_tokens.append({
            'committee_idx': ci,
            'token_number': token_num or str(r - 1),
            'customer_key': ckey,
        })
        
        # Collection records from date columns
        for col, date_str in date_columns:
            amount = clean_amount(ws.cell(r, col).value)
            if amount and amount > 0:
                all_collections.append({
                    'customer_key': ckey,
                    'committee_idx': ci,
                    'amount': amount,
                    'date': date_str,
                    'token_number': token_num or str(r - 1),
                })
        
        member_count += 1
    
    print(f"    Members extracted: {member_count}")

print(f"\n  Total unique customers (from committees): {len(all_customers)}")
print(f"  Total tokens: {len(all_tokens)}")
print(f"  Total collection records: {len(all_collections)}")

# ─── 2. Extract BYAJ KI LIST (Interest Accounts) ───────────────────────────

print("\n=== Phase 2: Extracting interest accounts (BYAJ KI LIST) ===")

interests = []
ws = wb['BYAJ KI LIST']
for r in range(2, ws.max_row + 1):
    name_val = clean_str(ws.cell(r, 1).value)
    if not name_val:
        continue
    
    ref_name = clean_str(ws.cell(r, 2).value)
    address = clean_str(ws.cell(r, 3).value)
    mobile = clean_mobile(ws.cell(r, 4).value)
    ref_mobile = clean_mobile(ws.cell(r, 5).value)
    interest_date = clean_str(ws.cell(r, 6).value)
    interest_amount = clean_amount(ws.cell(r, 7).value)
    reply = clean_str(ws.cell(r, 8).value)
    reason = clean_str(ws.cell(r, 9).value)
    
    # Also add to customer pool
    ckey = make_customer_key(name_val, mobile)
    if ckey not in all_customers:
        all_customers[ckey] = {
            'name': name_val,
            'mobile': mobile or 'N/A',
            'reference_name': ref_name,
            'reference_mobile': ref_mobile,
            'address': address,
            'reason': reason,
            'committees': [],
        }
    
    interests.append({
        'customer_key': ckey,
        'customer_name': name_val,
        'interest_date': interest_date,
        'interest_amount': interest_amount,
        'reply': reply,
        'reason': reason,
    })

print(f"  Interest accounts extracted: {len(interests)}")

# ─── 3. Extract Loans (nikku ji loan) ──────────────────────────────────────

print("\n=== Phase 3: Extracting loans (nikku ji loan) ===")

loans = []
ws = wb['nikku ji loan']
for r in range(2, ws.max_row + 1):
    name_val = clean_str(ws.cell(r, 1).value)
    if not name_val:
        continue
    
    date_val = date_to_str(ws.cell(r, 2).value)
    credit_cash = clean_amount(ws.cell(r, 3).value)
    credit_online = clean_amount(ws.cell(r, 4).value)
    debit_cash = clean_amount(ws.cell(r, 5).value)
    debit_online = clean_amount(ws.cell(r, 6).value)
    loan_amount = clean_amount(ws.cell(r, 7).value)
    loan_deposit = clean_amount(ws.cell(r, 8).value)
    interest = clean_amount(ws.cell(r, 9).value)
    
    mobile_val = None
    # Try to find mobile in later columns
    for c in range(10, min(ws.max_column + 1, 15)):
        v = clean_mobile(ws.cell(r, c).value)
        if v:
            mobile_val = v
            break
    
    ckey = make_customer_key(name_val, mobile_val)
    if ckey not in all_customers:
        all_customers[ckey] = {
            'name': name_val,
            'mobile': mobile_val or 'N/A',
            'reference_name': None,
            'reference_mobile': None,
            'address': None,
            'reason': None,
            'committees': [],
        }
    
    loans.append({
        'customer_key': ckey,
        'customer_name': name_val,
        'date': date_val,
        'credit_cash': credit_cash,
        'credit_online': credit_online,
        'debit_cash': debit_cash,
        'debit_online': debit_online,
        'loan_amount': loan_amount,
        'loan_deposit': loan_deposit,
        'interest': interest,
    })

print(f"  Loan records extracted: {len(loans)}")

# ─── 4. Extract Daily Collections ──────────────────────────────────────────

print("\n=== Phase 4: Extracting daily collections ===")

daily_collections = []
for sname in ['Daily collection', 'COLLECTION office 2', 'recovery collection']:
    ws = wb[sname]
    print(f"  Processing: {sname}")
    count = 0
    for r in range(2, ws.max_row + 1):
        name_val = clean_str(ws.cell(r, 1).value)
        if not name_val:
            continue
        
        date_val = date_to_str(ws.cell(r, 2).value)
        credit_cash = clean_amount(ws.cell(r, 3).value)
        credit_online = clean_amount(ws.cell(r, 4).value)
        debit_cash = clean_amount(ws.cell(r, 5).value)
        debit_online = clean_amount(ws.cell(r, 6).value)
        debit_gift = clean_amount(ws.cell(r, 7).value)
        
        # Token references in cols 8-10 (5th date, 15th date, 20th date bissi)
        token_5th = clean_str(ws.cell(r, 8).value)
        token_15th = clean_str(ws.cell(r, 9).value)
        token_20th = clean_str(ws.cell(r, 10).value)
        
        ckey = make_customer_key(name_val, None)
        if ckey not in all_customers:
            all_customers[ckey] = {
                'name': name_val,
                'mobile': 'N/A',
                'reference_name': None,
                'reference_mobile': None,
                'address': None,
                'reason': None,
                'committees': [],
            }
        
        daily_collections.append({
            'customer_key': ckey,
            'customer_name': name_val,
            'date': date_val,
            'credit_cash': credit_cash,
            'credit_online': credit_online,
            'debit_cash': debit_cash,
            'debit_online': debit_online,
            'debit_gift': debit_gift,
            'token_5th': token_5th,
            'token_15th': token_15th,
            'token_20th': token_20th,
            'source': sname,
        })
        count += 1
    print(f"    Records: {count}")

print(f"  Total daily collection records: {len(daily_collections)}")

# ─── 5. Extract Gift Records ────────────────────────────────────────────────

print("\n=== Phase 5: Extracting gift records ===")

GIFT_RECORD_SHEETS = [
    'Sawariya seth bissi gift record',
    'Pyare mohan bissi gift records',
    'Hare ka sahara bissi gift recor',
    'Shree krishna aasociates gift r',
]

# Map gift record sheets to committee names
GIFT_COMMITTEE_MAP = {
    'Sawariya seth bissi gift record': 'Sawariya Seth Bissi',
    'Pyare mohan bissi gift records': 'Pyare Mohan Bissi',
    'Hare ka sahara bissi gift recor': 'Hare Ka Sahara Bissi',
    'Shree krishna aasociates gift r': 'Shree Krishna Associates',
}

gifts = []
for sname in GIFT_RECORD_SHEETS:
    ws = wb[sname]
    print(f"  Processing: {sname}")
    
    # Find date columns
    date_cols = []
    for c in range(5, ws.max_column + 1):
        d = date_to_str(ws.cell(1, c).value)
        if d:
            date_cols.append((c, d))
    
    count = 0
    for r in range(2, ws.max_row + 1):
        token_val = clean_str(ws.cell(r, 1).value)
        name_val = clean_str(ws.cell(r, 2).value)
        if not name_val and not token_val:
            continue
        
        ref_name = clean_str(ws.cell(r, 3).value)
        mobile = clean_mobile(ws.cell(r, 4).value)
        
        ckey = make_customer_key(name_val or f"Token-{token_val}", mobile)
        
        # Check each date column for gift
        for col, date_str in date_cols:
            gift_val = clean_str(ws.cell(r, col).value)
            if gift_val and gift_val.lower() not in ('none', 'nan'):
                gifts.append({
                    'customer_key': ckey,
                    'customer_name': name_val,
                    'token_number': token_val,
                    'committee_name': GIFT_COMMITTEE_MAP.get(sname, sname),
                    'date': date_str,
                    'gift_name': gift_val,
                })
                count += 1
    print(f"    Gift records: {count}")

print(f"  Total gift records: {len(gifts)}")

# ─── 6. Extract Lucky Draw (Lucky Token list + gift sheets) ─────────────────

print("\n=== Phase 6: Extracting lucky draw data ===")

GIFT_SHEET_NAMES = [
    'Sawariya bissi 5 date gift shee',
    'Pyare Mohan bissi gift sheets',
    'Hare ka sahara bissi gift sheet',
    'Shree krishna gift sheet',
]

GIFT_SHEET_COMMITTEE_MAP = {
    'Sawariya bissi 5 date gift shee': 'Sawariya Seth Bissi',
    'Pyare Mohan bissi gift sheets': 'Pyare Mohan Bissi',
    'Hare ka sahara bissi gift sheet': 'Hare Ka Sahara Bissi',
    'Shree krishna gift sheet': 'Shree Krishna Associates',
}

lotteries = []
for sname in GIFT_SHEET_NAMES:
    ws = wb[sname]
    print(f"  Processing: {sname}")
    
    # These sheets have a different layout:
    # Row 1: Date headers (grouped in sets of 4 cols: Name, Token, Gift Status, blank)
    # Row 2: Column headers (Name, Token No, Gift Status, blank, repeated)
    # Rows 3+: Data
    
    # Find date groups from row 1
    date_groups = []
    for c in range(1, ws.max_column + 1):
        d = date_to_str(ws.cell(1, c).value)
        if d:
            date_groups.append({'date': d, 'start_col': c})
    
    count = 0
    for dg in date_groups:
        sc = dg['start_col']
        date_str = dg['date']
        
        # In each group: col+0=Name, col+1=Token, col+2=Gift Status
        for r in range(3, ws.max_row + 1):
            name_val = clean_str(ws.cell(r, sc).value)
            token_val = clean_str(ws.cell(r, sc + 1).value)
            status_val = clean_str(ws.cell(r, sc + 2).value)
            
            if not name_val and not token_val:
                continue
            
            if token_val:
                token_val = re.sub(r'\.0$', '', token_val)
            
            lotteries.append({
                'committee_name': GIFT_SHEET_COMMITTEE_MAP.get(sname, sname),
                'date': date_str,
                'winner_name': name_val,
                'token_number': token_val,
                'gift_status': status_val,
            })
            count += 1
    
    print(f"    Lucky draw entries: {count}")

print(f"  Total lucky draw entries: {len(lotteries)}")

# ─── 7. Also check Manager collection + monthly installment ────────────────

print("\n=== Phase 7: Extracting manager & monthly installment data ===")

for sname in ['Manager collection', 'nikku ji online']:
    if sname in wb.sheetnames:
        ws = wb[sname]
        print(f"  Processing: {sname}")
        count = 0
        for r in range(2, ws.max_row + 1):
            name_val = clean_str(ws.cell(r, 1).value)
            if not name_val:
                continue
            date_val = date_to_str(ws.cell(r, 2).value)
            credit_cash = clean_amount(ws.cell(r, 3).value)
            credit_online = clean_amount(ws.cell(r, 4).value)
            debit_cash = clean_amount(ws.cell(r, 5).value)
            debit_online = clean_amount(ws.cell(r, 6).value)
            
            ckey = make_customer_key(name_val, None)
            if ckey not in all_customers:
                all_customers[ckey] = {
                    'name': name_val,
                    'mobile': 'N/A',
                    'reference_name': None,
                    'reference_mobile': None,
                    'address': None,
                    'reason': None,
                    'committees': [],
                }
            
            daily_collections.append({
                'customer_key': ckey,
                'customer_name': name_val,
                'date': date_val,
                'credit_cash': credit_cash,
                'credit_online': credit_online,
                'debit_cash': debit_cash,
                'debit_online': debit_online,
                'debit_gift': None,
                'token_5th': None,
                'token_15th': None,
                'token_20th': None,
                'source': sname,
            })
            count += 1
        print(f"    Records: {count}")

# ─── Final: Save all JSON ───────────────────────────────────────────────────

print(f"\n=== Final Stats ===")
print(f"  Total unique customers: {len(all_customers)}")
print(f"  Committees: {len(committees_data)}")
print(f"  Tokens: {len(all_tokens)}")
print(f"  Committee collections: {len(all_collections)}")
print(f"  Interest accounts: {len(interests)}")
print(f"  Loan records: {len(loans)}")
print(f"  Daily collections: {len(daily_collections)}")
print(f"  Gift records: {len(gifts)}")
print(f"  Lottery entries: {len(lotteries)}")

# Convert customers dict to list with index
customer_list = []
ckey_to_idx = {}
for i, (ckey, cdata) in enumerate(all_customers.items()):
    cdata['_key'] = ckey
    cdata['_idx'] = i
    customer_list.append(cdata)
    ckey_to_idx[ckey] = i

# Save all
def save_json(name, data):
    path = os.path.join(OUTPUT_DIR, name)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    print(f"  Saved {path} ({len(data)} records)")

save_json('extracted_customers.json', customer_list)
save_json('extracted_committees.json', committees_data)
save_json('extracted_tokens.json', all_tokens)
save_json('extracted_collections.json', all_collections)
save_json('extracted_interests.json', interests)
save_json('extracted_loans.json', loans)
save_json('extracted_daily_collections.json', daily_collections)
save_json('extracted_gifts.json', gifts)
save_json('extracted_lotteries.json', lotteries)

print("\n✅ All data extracted successfully!")

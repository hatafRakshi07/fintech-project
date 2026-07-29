import os
import re
import uuid
import openpyxl
import psycopg2
import psycopg2.extras
from datetime import datetime

SUPABASE_URL = "postgresql://postgres:hatafrakshi@db.qnflaeexcmwwcabrcrhb.supabase.co:5432/postgres"
WORKBOOK_PATH = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder.xlsx"

def clean_mobile(val):
    if not val:
        return None
    s = re.sub(r"[^\d]", "", str(val).split(".")[0])
    s = s[-10:] if len(s) >= 10 else s
    return s if len(s) >= 6 else None

def clean_name(val):
    if not val:
        return None
    v = str(val).strip()
    return v[:200] if v and v.lower() not in ("none", "name", "name ", "", "jsk") else None

print("Connecting to Supabase PostgreSQL Database for Drizzle ERP Schema Sync...")
conn = psycopg2.connect(SUPABASE_URL, sslmode="require")
cur = conn.cursor()

print("Fast-loading Bissi folder.xlsx workbook (read-only)...")
wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True)

# Truncate Drizzle Schema tables
cur.execute("""
    TRUNCATE TABLE payment_items, payment_receipts, memberships, schemes CASCADE;
""")
conn.commit()

schemes_config = [
    {
        "name": "Sawariya Seth Bissi (5th Date)",
        "code": "SSB5",
        "main_sheet": "Sawariya seth 5 date",
        "amount": 3000.0,
        "draw_day": 5,
        "member_limit": 500
    },
    {
        "name": "Pyare Mohan Bissi (15th Date)",
        "code": "PMB15",
        "main_sheet": "Pyare mohan 15 date",
        "amount": 3000.0,
        "draw_day": 15,
        "member_limit": 500
    },
    {
        "name": "Hare Ka Sahara Bissi (20th Date)",
        "code": "HKS20",
        "main_sheet": "Hare ka sahara bissi 20 date",
        "amount": 2500.0,
        "draw_day": 20,
        "member_limit": 500
    },
    {
        "name": "Shree Krishna Associate Bissi",
        "code": "SKA26",
        "main_sheet": "Shree Krishna associate lottery",
        "amount": 3000.0,
        "draw_day": 26,
        "member_limit": 1111
    }
]

# Create Schemes in Drizzle 'schemes' table
for sc in schemes_config:
    scheme_id = str(uuid.uuid4())
    cur.execute("""
        INSERT INTO schemes (
            id, name, code, draw_day, draw_time, start_date, monthly_installment, duration_months, status, created_at, updated_at
        ) VALUES (
            %s, %s, %s, %s, '12:00:00', '2025-11-05', %s, 20, 'ACTIVE', NOW(), NOW()
        )
    """, (scheme_id, sc["name"], sc["code"], sc["draw_day"], sc["amount"]))
    sc["id"] = scheme_id

conn.commit()

receipt_counter = 1

for sc in schemes_config:
    c_name = sc["name"]
    s_id = sc["id"]
    main_sh = sc["main_sheet"]
    mem_limit = sc["member_limit"]
    
    if main_sh not in wb.sheetnames:
        continue
    
    print(f"[Syncing Drizzle Schema] {c_name} (Sheet: {main_sh})...")
    ws = wb[main_sh]
    rows = [r for r in ws.iter_rows(values_only=True) if r and any(c is not None for c in r)]
    if not rows:
        continue

    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    token_idx, name_idx, ref_name_idx, mobile_idx = 0, 1, 2, 3
    for i, h in enumerate(headers):
        if "token" in h: token_idx = i
        elif "reference name" in h: ref_name_idx = i
        elif "name" in h and "reference" not in h: name_idx = i
        elif "contact" in h or ("mobile" in h and "reference" not in h and "no" in h): mobile_idx = i

    processed_tokens = set()
    memberships_batch = []
    receipts_batch = []
    items_batch = []

    for r_idx in range(1, len(rows)):
        row = rows[r_idx]
        if not row: continue
        raw_token_val = row[token_idx] if token_idx < len(row) else None
        if not raw_token_val: continue
        raw_token = str(raw_token_val).strip().split(".")[0]
        if not raw_token or not raw_token.isdigit(): continue

        raw_name = clean_name(row[name_idx]) if name_idx < len(row) else None
        raw_mobile = clean_mobile(row[mobile_idx]) if mobile_idx < len(row) else None

        phone = raw_mobile or f"999{int(raw_token):07d}"
        cust_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"bissi_cust_{phone}"))
        mem_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"mem_{s_id}_{raw_token}"))

        memberships_batch.append((mem_uuid, cust_uuid, s_id, '2025-11-05', 'ACTIVE', datetime.now(), datetime.now()))
        processed_tokens.add(raw_token)

        # Parse collections into paymentReceipts & paymentItems
        for c_i in range(mobile_idx + 1, len(row)):
            cell_val = str(row[c_i]).strip() if row[c_i] is not None else ""
            if cell_val and cell_val.lower() not in ("none", "-", "name"):
                clean_num = re.sub(r"[^\d.]", "", cell_val)
                try:
                    amt = float(clean_num) if clean_num and clean_num != "." else sc["amount"]
                except Exception:
                    amt = sc["amount"]

                if 0 < amt <= 100000:
                    rcpt_uuid = str(uuid.uuid4())
                    rcpt_no = f"REC-{receipt_counter:07d}"
                    receipt_counter += 1

                    receipts_batch.append((rcpt_uuid, rcpt_no, cust_uuid, 'CASH', amt, f"{c_name} Token #{raw_token}", datetime.now(), datetime.now()))
                    items_batch.append((str(uuid.uuid4()), rcpt_uuid, 'INSTALLMENT', amt, mem_uuid, datetime.now(), datetime.now()))

    # Fill remaining seats
    for t_num in range(1, mem_limit + 1):
        str_t = str(t_num)
        if str_t not in processed_tokens:
            phone = f"999{t_num:07d}"
            cust_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"bissi_cust_{phone}"))
            mem_uuid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f"mem_{s_id}_{str_t}"))
            memberships_batch.append((mem_uuid, cust_uuid, s_id, '2025-11-05', 'ACTIVE', datetime.now(), datetime.now()))

    if memberships_batch:
        psycopg2.extras.execute_values(
            cur,
            """
            INSERT INTO memberships (id, customer_id, scheme_id, joining_date, status, created_at, updated_at)
            VALUES %s
            ON CONFLICT (id) DO NOTHING
            """,
            memberships_batch
        )

    if receipts_batch:
        psycopg2.extras.execute_values(
            cur,
            """
            INSERT INTO payment_receipts (id, receipt_no, customer_id, payment_method, total_amount, notes, created_at, updated_at)
            VALUES %s
            ON CONFLICT (id) DO NOTHING
            """,
            receipts_batch
        )

    if items_batch:
        psycopg2.extras.execute_values(
            cur,
            """
            INSERT INTO payment_items (id, receipt_id, type, amount, reference_id, created_at, updated_at)
            VALUES %s
            ON CONFLICT (id) DO NOTHING
            """,
            items_batch
        )

    conn.commit()
    print(f"  * Drizzle Schema for {c_name} synced! ({len(receipts_batch)} Payment Receipts, {len(memberships_batch)} Memberships)")

print("\n=======================================================")
print("  DRIZZLE SCHEMA BACKWARD COMPATIBILITY SYNC COMPLETE!")
print("=======================================================")

cur.execute("SELECT COUNT(*) FROM schemes WHERE status = 'ACTIVE'")
print("Active schemes in Drizzle:", cur.fetchone()[0])

cur.execute("SELECT COUNT(*) FROM memberships")
print("Total Memberships in Drizzle:", cur.fetchone()[0])

cur.execute("SELECT COUNT(*) FROM payment_receipts")
print("Total Payment Receipts in Drizzle:", cur.fetchone()[0])

cur.execute("SELECT COUNT(*) FROM payment_items")
print("Total Payment Items in Drizzle:", cur.fetchone()[0])

cur.close()
conn.close()

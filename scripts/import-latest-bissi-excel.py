import os
import re
import sys
import openpyxl
import psycopg2
from datetime import datetime, date

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

NEON_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require"
)
EXCEL_PATH = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder (1).xlsx"

def clean_phone(val):
    if not val: return None
    s = re.sub(r"[^\d]", "", str(val).split(".")[0])
    s = s[-10:] if len(s) >= 10 else s
    return s if len(s) >= 6 else None

def clean_name(val):
    if not val: return None
    v = str(val).strip()
    v = re.sub(r"\s+", " ", v)
    return v if v and v.lower() not in ("none", "name", "name ", "", "jsk", "-") else None

def parse_date(val, default_date=None):
    if not val:
        return default_date or datetime.now()
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    
    s = str(val).strip()
    parts = re.split(r"[/-]", s)
    if len(parts) == 3:
        try:
            d = int(parts[0])
            m = int(parts[1])
            y = int(parts[2])
            if y < 100: y += 2000
            return datetime(y, m, d)
        except Exception:
            pass
    return default_date or datetime.now()

def run_import():
    print("=== EXCEL DATA IMPORT WITH ACCURATE DATES ===")
    print(f"Connecting to database: {NEON_URL[:45]}...")
    conn = psycopg2.connect(NEON_URL)
    cur = conn.cursor()

    # Ensure Branch
    cur.execute("""
        INSERT INTO branches (id, name, code, city, address, status, updated_at)
        VALUES (1, 'Shree Krishna Associate', 'SKA001', 'Main City', 'Main Branch', 'active', NOW())
        ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name
    """)

    # Ensure 4 Bissi Committees
    committees_def = [
        (1, 'Sawariya Seth Bissi', 'monthly', 3000.0, 500),
        (2, 'Pyare Mohan Bissi', 'monthly', 3000.0, 500),
        (3, 'Hare Ka Sahara Bissi', 'monthly', 3000.0, 500),
        (4, 'Shree Krishna Bissi', 'monthly', 3000.0, 1111),
    ]
    for c in committees_def:
        cur.execute("""
            INSERT INTO committees (id, name, type, installment_amount, member_limit, branch_id, status, updated_at)
            VALUES (%s, %s, %s, %s, %s, 1, 'active', NOW())
            ON CONFLICT (id) DO UPDATE 
            SET name = EXCLUDED.name, installment_amount = EXCLUDED.installment_amount, member_limit = EXCLUDED.member_limit, updated_at = NOW()
        """, c)
    conn.commit()

    print("Loading workbook into memory...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Loaded workbook with {len(wb.sheetnames)} sheets.")

    # Memory cache for customers
    cur.execute("SELECT id, name, mobile FROM customers")
    customers = {}
    for cid, name, mob in cur.fetchall():
        if mob: customers[mob] = cid
        if name: customers[name.lower().strip()] = cid

    def get_or_create_customer(name, mobile=None, address=None):
        name_clean = clean_name(name)
        mob_clean = clean_phone(mobile)
        
        if mob_clean and mob_clean in customers:
            return customers[mob_clean]
        if name_clean and name_clean.lower() in customers:
            return customers[name_clean.lower()]

        final_name = name_clean or f"Member #{mob_clean or 'Unknown'}"
        final_mob = mob_clean or f"9000{len(customers)+100000:06d}"
        ref_no = f"REF-{len(customers) + 1001}"
        
        cur.execute("""
            INSERT INTO customers (name, mobile, reference_number, address, branch_id, status, created_at, updated_at)
            VALUES (%s, %s, %s, %s, 1, 'active', NOW(), NOW())
            RETURNING id
        """, (final_name, final_mob, ref_no, address))
        cid = cur.fetchone()[0]
        customers[final_mob] = cid
        if name_clean: customers[name_clean.lower()] = cid
        return cid

    collections_count = 0
    installments_count = 0

    # 1. Process "Daily collection" sheet
    if "Daily collection" in wb.sheetnames:
        print("\nProcessing 'Daily collection' sheet...")
        ws = wb["Daily collection"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            header = [str(c).strip() if c else "" for c in rows[0]]
            for row in rows[1:]:
                if not row or not any(row): continue
                c_name = row[0]
                dt_val = row[1]
                credit_cash = float(row[2]) if row[2] and str(row[2]).replace(".","").isdigit() else 0.0
                credit_online = float(row[3]) if len(row) > 3 and row[3] and str(row[3]).replace(".","").isdigit() else 0.0
                
                if not c_name and credit_cash == 0 and credit_online == 0: continue
                
                cust_id = get_or_create_customer(c_name or "Daily Collector")
                exact_date = parse_date(dt_val)
                amount = credit_cash + credit_online
                pmode = "bank" if credit_online > 0 else "cash"

                if amount > 0:
                    receipt = f"REC-DC-{exact_date.strftime('%Y%m%d')}-{collections_count + 1000}"
                    notes = f"Daily collection entry for {clean_name(c_name) or 'Member'}"
                    cur.execute("""
                        INSERT INTO collections (customer_id, collector_id, branch_id, amount, payment_mode, receipt_number, notes, verification_status, collected_at, created_at)
                        VALUES (%s, 1, 1, %s, %s, %s, %s, 'verified', %s, %s)
                    """, (cust_id, amount, pmode, receipt, notes, exact_date, exact_date))
                    collections_count += 1

    # 2. Process Lottery/Bissi and Monthly Installment sheets for Installments with exact dates (EXCLUDING LOAN SHEETS)
    bissi_sheets = [
        ("Shree Krishna associate lottery", 4),
        ("Shree krishna aasociates gift r", 4),
        ("Radhe krishna bissi gift list", 4),
        ("MONTHLY INSTALLMENT", 1),
        (" monthly payment details", 1),
    ]

    for sheet_name, comm_id in bissi_sheets:
        if "loan" in sheet_name.lower():
            print(f"\nSkipping loan sheet: '{sheet_name}'")
            continue

        if sheet_name in wb.sheetnames:
            print(f"\nProcessing '{sheet_name}' sheet...")
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                header = rows[0]
                # Collect date headers from column index 5+
                date_cols = []
                for idx, col in enumerate(header):
                    if idx >= 5 and col:
                        dt = parse_date(col, None)
                        if dt: date_cols.append((idx, dt))
                
                for row in rows[1:]:
                    if not row or not row[0]: continue
                    tok_no = str(row[0]).split(".")[0].strip()
                    c_name = row[1] if len(row) > 1 else None
                    mob = row[3] if len(row) > 3 else None
                    
                    if not tok_no.isdigit(): continue
                    token_num = int(tok_no)
                    cust_id = get_or_create_customer(c_name or f"Token #{token_num}", mob)
                    
                    # Ensure Token & Member
                    cur.execute("""
                        INSERT INTO tokens (token_number, committee_id, customer_id, status, created_at)
                        VALUES (%s, %s, %s, 'ACTIVE', NOW())
                        ON CONFLICT DO NOTHING
                    """, (str(token_num), comm_id, cust_id))
                    
                    cur.execute("""
                        INSERT INTO committee_members (committee_id, customer_id, token_number, status, joined_at)
                        VALUES (%s, %s, %s, 'active', NOW())
                        ON CONFLICT DO NOTHING
                    """, (comm_id, cust_id, token_num))

                    # Process monthly payments across date columns
                    for col_idx, exact_date in date_cols:
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val and (isinstance(val, (int, float)) or str(val).replace(".","").isdigit()):
                                amt = float(val)
                                if amt > 0:
                                    rec = f"REC-INSTALL-{token_num}-{exact_date.strftime('%Y%m')}"
                                    cur.execute("""
                                        INSERT INTO installments (customer_id, collector_id, token_id, committee_id, month, year, amount, payment_date, payment_mode, receipt_number, created_at)
                                        SELECT %s, 1, id, %s, %s, %s, %s, %s, 'cash', %s, %s
                                        FROM tokens WHERE token_number = %s AND committee_id = %s LIMIT 1
                                    """, (cust_id, comm_id, exact_date.month, exact_date.year, amt, exact_date, rec, exact_date, token_num, comm_id))
                                    installments_count += 1

    conn.commit()
    print(f"\nImport completed successfully!")
    print(f"Total Collections imported: {collections_count}")
    print(f"Total Installments imported: {installments_count}")
    cur.close()
    conn.close()

if __name__ == "__main__":
    run_import()

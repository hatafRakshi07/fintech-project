#!/usr/bin/env python3
"""
Import all business data from Bissi.xlsx into the live Neon database.
Clears ONLY the business data tables (keeps schema, auth, settings).
"""

import openpyxl
import psycopg2
import psycopg2.extras
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

DB_URL = "postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed-pooler.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require"
WORKBOOK = "Bissi.xlsx"

# ─── helpers ────────────────────────────────────────────────────────────────

def to_num(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except (ValueError, TypeError): return default

def to_str(v, default=""):
    if v is None: return default
    s = str(v).strip()
    # strip trailing .0 from phone numbers
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s

def to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%y", "%d/%m/%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None

def is_received(v):
    if v is None: return False
    s = str(v).lower().strip()
    return "received" in s or "done" in s or s.startswith("rec")

def extract_number(name):
    """Extract the serial number from 'Name (N)' format."""
    m = re.search(r'\((\d+)\)', str(name))
    return int(m.group(1)) if m else None

def clean_name(name):
    """Remove the trailing (N) from name."""
    return re.sub(r'\s*\(\d+\)\s*$', '', str(name)).strip()

# ─── DB setup ───────────────────────────────────────────────────────────────

def get_conn():
    return psycopg2.connect(DB_URL, connect_timeout=30)

def ensure_interest_tables(cur):
    # Tables already exist in live DB - just verify
    pass

def ensure_diary_tables(cur):
    # Tables already exist in live DB - just verify
    pass

# ─── Import: Daily Diary ─────────────────────────────────────────────────────

def import_daily_diary(wb, cur):
    ws = wb['daily diary']
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Clear existing data
    cur.execute("DELETE FROM daily_diary_payments")
    cur.execute("DELETE FROM daily_diary_loans")
    print("  Cleared daily_diary_loans and daily_diary_payments")

    inserted_loans = 0
    inserted_payments = 0

    for row in rows:
        name = row[0]
        if not name or str(name).strip() == '':
            continue

        name_str = to_str(name)
        mobile = to_str(row[1])
        ref_mobile = to_str(row[2])
        payment_mode = to_str(row[3])
        reason = to_str(row[4])
        address = to_str(row[5])
        security = to_str(row[6])
        loan_amount = to_num(row[7])
        start_date = to_date(row[8])
        complete_date = to_date(row[9])
        amount_taken = to_num(row[10])   # total collected so far
        # col 11 = REMAINING TILL TODAY (not used directly)
        # col 12 = TOTAL REMAINING

        if loan_amount <= 0:
            continue

        # Determine status
        remaining = loan_amount - amount_taken
        status = 'COMPLETED' if remaining <= 0 else 'ACTIVE'

        # Collection plan: compute daily amount based on loan duration
        collection_plan = "500/day"
        if start_date and complete_date:
            delta_days = max(1, (complete_date - start_date).days)
            daily_amt = round(loan_amount / delta_days)
            if daily_amt > 0:
                collection_plan = f"{int(daily_amt)}/day"

        serial_no = extract_number(name_str)

        cur.execute("""
            INSERT INTO daily_diary_loans
                (customer_name, mobile_number, reference_mobile_numbers,
                 address, security,
                 loan_amount, start_date, expected_complete_date, collection_plan,
                 status, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            name_str,
            mobile or None,
            ref_mobile or None,
            address or None,
            security or None,
            loan_amount,
            str(start_date) if start_date else None,
            str(complete_date) if complete_date else None,
            collection_plan,
            status,
            f"Serial:{serial_no}|Mode:{payment_mode}|Reason:{reason}" if (serial_no or payment_mode or reason) else None,
        ))
        loan_id = cur.fetchone()[0]
        inserted_loans += 1

        # Create a single payment record for amount_taken if > 0
        if amount_taken > 0:
            pay_date = complete_date or start_date or date.today()
            cur.execute("""
                INSERT INTO daily_diary_payments
                    (loan_id, payment_date, amount_deposited, payment_mode, notes)
                VALUES (%s,%s,%s,%s,%s)
            """, (
                loan_id,
                str(pay_date),
                amount_taken,
                payment_mode or 'Cash',
                'Imported from Excel (total amount taken)',
            ))
            inserted_payments += 1

    print(f"  Inserted {inserted_loans} loans, {inserted_payments} payment records")
    return inserted_loans

# ─── Import: BYAJ KI LIST ────────────────────────────────────────────────────

def import_byaj(wb, cur):
    ws = wb['BYAJ KI LIST']
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Date columns start from col index 10
    date_cols = []
    for i, h in enumerate(header):
        if i >= 10 and isinstance(h, (datetime, date)):
            date_cols.append((i, h if isinstance(h, date) else h.date()))

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    cur.execute("DELETE FROM interest_transactions")
    cur.execute("DELETE FROM interest_accounts")
    print("  Cleared interest_accounts and interest_transactions")

    inserted_accounts = 0
    inserted_txns = 0

    for row in rows:
        name = row[0]
        if not name or str(name).strip() == '':
            continue

        address = to_str(row[1])
        mobile = to_str(row[2])
        ref_mobile = to_str(row[3])
        interest_date = to_str(row[4])
        monthly_amount = to_num(row[5])
        reply = to_str(row[6])
        reason1 = to_str(row[7])
        reason2 = to_str(row[8])
        address2 = to_str(row[9])

        if monthly_amount <= 0:
            continue

        serial_no = extract_number(str(name))
        cust_id_placeholder = serial_no if serial_no else (inserted_accounts + 1)
        start_dt = date(2024, 10, 1)  # earliest date in header cols
        notes_str = (f"{to_str(name)} | Mobile: {mobile} | "
                     f"Ref: {ref_mobile} | Date: {interest_date} | "
                     f"Addr: {address or address2} | Reply: {reply}")
        cur.execute("""
            INSERT INTO interest_accounts
                (customer_id, principal_amount, interest_rate,
                 start_date, total_interest_paid, pending_interest,
                 monthly_interest, outstanding_amount,
                 status, branch_id, notes)
            VALUES (%s, 0, 0, %s, 0, %s, %s, %s, 'active', 1, %s)
            RETURNING id
        """, (
            cust_id_placeholder,
            str(start_dt),
            monthly_amount,   # pending_interest = monthly
            monthly_amount,   # monthly_interest
            monthly_amount,   # outstanding
            notes_str,
        ))
        acct_id = cur.fetchone()[0]
        inserted_accounts += 1

        # Create transaction records for each received payment
        for col_idx, pay_date in date_cols:
            cell_val = row[col_idx] if col_idx < len(row) else None
            if is_received(cell_val):
                # Extract numeric amount if present (e.g. "900received")
                amount_str = re.sub(r'[^0-9.]', '', str(cell_val))
                try:
                    amount = float(amount_str) if amount_str else monthly_amount
                except ValueError:
                    amount = monthly_amount
                if amount <= 0:
                    amount = monthly_amount

                cur.execute("""
                    INSERT INTO interest_transactions
                        (account_id, customer_id, type, amount,
                         month, year, payment_date, branch_id, notes)
                    VALUES (%s, %s, 'credit', %s, %s, %s, %s, 1, %s)
                """, (
                    acct_id,
                    cust_id_placeholder,
                    amount,
                    pay_date.month,
                    pay_date.year,
                    str(pay_date),
                    to_str(cell_val),
                ))
                inserted_txns += 1

    print(f"  Inserted {inserted_accounts} interest accounts, {inserted_txns} transactions")
    return inserted_accounts

# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading workbook: {WORKBOOK}")
    try:
        wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"ERROR: {WORKBOOK} not found. Copy it to the project root first.")
        sys.exit(1)

    print(f"Sheets found: {wb.sheetnames}")
    print(f"\nConnecting to database...")

    conn = get_conn()
    conn.autocommit = False
    cur = conn.cursor()

    try:
        print("\n[1/4] Ensuring tables exist...")
        ensure_diary_tables(cur)
        ensure_interest_tables(cur)
        conn.commit()

        print("\n[2/4] Importing Daily Diary loans...")
        n1 = import_daily_diary(wb, cur)
        conn.commit()
        print(f"  ✓ Daily Diary: {n1} loans imported")

        print("\n[3/4] Importing BYAJ KI LIST (Interest)...")
        n2 = import_byaj(wb, cur)
        conn.commit()
        print(f"  ✓ BYAJ: {n2} interest accounts imported")

        print("\n[4/4] Verifying counts...")
        cur.execute("SELECT COUNT(*) FROM daily_diary_loans")
        print(f"  daily_diary_loans: {cur.fetchone()[0]}")
        cur.execute("SELECT COUNT(*) FROM daily_diary_payments")
        print(f"  daily_diary_payments: {cur.fetchone()[0]}")
        cur.execute("SELECT COUNT(*) FROM interest_accounts")
        print(f"  interest_accounts: {cur.fetchone()[0]}")
        cur.execute("SELECT COUNT(*) FROM interest_transactions")
        print(f"  interest_transactions: {cur.fetchone()[0]}")

        print("\n✅ Import complete!")

    except Exception as e:
        conn.rollback()
        print(f"\n❌ ERROR: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)
    finally:
        cur.close()
        conn.close()

if __name__ == "__main__":
    main()

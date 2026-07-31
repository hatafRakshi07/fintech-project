import os
import re
import sys
import openpyxl
import psycopg2
import psycopg2.extras
from datetime import datetime, date

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

NEON_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require"
)
EXCEL_PATH = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder (1).xlsx"

def clean_phone(val):
    if not val: return None
    s = re.sub(r"[^\d]", "", str(val).split(".")[0])
    s = s[-10:] if len(s) >= 10 else s
    return s if len(s) >= 6 else None

def clean_name(val):
    if not val: return None
    v = str(val).strip()
    v = re.sub(r"\s+", " ", v)
    return v if v and v.lower() not in ("none", "name", "name ", "", "jsk", "-") else None

def parse_date(val, default_date=None):
    if not val: return default_date or datetime.now()
    if isinstance(val, datetime): return val
    if isinstance(val, date): return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    parts = re.split(r"[/-]", s)
    if len(parts) == 3:
        try:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100: y += 2000
            return datetime(y, m, d)
        except Exception: pass
    return default_date or datetime.now()

def run_fast_batch_import():
    print("=== HIGH SPEED BATCH IMPORT WITH ACCURATE DATES ===")
    conn = psycopg2.connect(NEON_URL)
    cur = conn.cursor()

    # 1. Update Hare Ka Sahara Bissi Installment Amount to 2500
    cur.execute("UPDATE committees SET installment_amount = 2500.0 WHERE id = 3 OR name LIKE '%Hare Ka Sahara%'")
    conn.commit()

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    cur.execute("SELECT id, name, mobile FROM customers")
    customers = {}
    for cid, name, mob in cur.fetchall():
        if mob: customers[mob] = cid
        if name: customers[name.lower().strip()] = cid

    def get_or_create_customer(name, mobile=None):
        name_clean = clean_name(name)
        mob_clean = clean_phone(mobile)
        if mob_clean and mob_clean in customers: return customers[mob_clean]
        if name_clean and name_clean.lower() in customers: return customers[name_clean.lower()]

        final_name = name_clean or f"Member #{mob_clean or 'Unknown'}"
        final_mob = mob_clean or f"9000{len(customers)+100000:06d}"
        ref_no = f"REF-{len(customers) + 1001}"
        cur.execute("""
            INSERT INTO customers (name, mobile, reference_number, branch_id, status, created_at, updated_at)
            VALUES (%s, %s, %s, 1, 'active', NOW(), NOW())
            RETURNING id
        """, (final_name, final_mob, ref_no))
        cid = cur.fetchone()[0]
        customers[final_mob] = cid
        if name_clean: customers[name_clean.lower()] = cid
        return cid

    collections_batch = []
    installments_batch = []
    lotteries_batch = []

    # Process Daily collection
    if "Daily collection" in wb.sheetnames:
        print("Batching 'Daily collection' sheet...")
        ws = wb["Daily collection"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            for row in rows[1:]:
                if not row or not any(row): continue
                c_name, dt_val = row[0], row[1]
                credit_cash = float(row[2]) if row[2] and str(row[2]).replace(".","").isdigit() else 0.0
                credit_online = float(row[3]) if len(row) > 3 and row[3] and str(row[3]).replace(".","").isdigit() else 0.0
                if not c_name and credit_cash == 0 and credit_online == 0: continue

                cust_id = get_or_create_customer(c_name or "Daily Collector")
                exact_date = parse_date(dt_val)
                amount = credit_cash + credit_online
                pmode = "bank" if credit_online > 0 else "cash"

                if amount > 0:
                    receipt = f"REC-DC-{exact_date.strftime('%Y%m%d')}-{len(collections_batch) + 1000}"
                    notes = f"Daily collection entry for {clean_name(c_name) or 'Member'}"
                    collections_batch.append((cust_id, 1, 1, amount, pmode, receipt, notes, 'verified', exact_date, exact_date))

    # Process Lottery & Gift sheets
    bissi_sheets = [
        ("Shree Krishna associate lottery", 4),
        ("Shree krishna aasociates gift r", 4),
        ("Radhe krishna bissi gift list", 4),
        ("MONTHLY INSTALLMENT", 1),
        (" monthly payment details", 1),
    ]

    for sheet_name, comm_id in bissi_sheets:
        if "loan" in sheet_name.lower(): continue
        if sheet_name in wb.sheetnames:
            print(f"Batching '{sheet_name}' sheet...")
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                header = rows[0]
                date_cols = []
                for idx, col in enumerate(header):
                    if idx >= 1 and col:
                        dt = parse_date(col, None)
                        if dt: date_cols.append((idx, dt))

                for row in rows[1:]:
                    if not row or not row[0]: continue
                    tok_no = str(row[0]).split(".")[0].strip()
                    if not tok_no.isdigit(): continue
                    token_num = int(tok_no)
                    c_name = row[1] if len(row) > 1 else None
                    mob = row[3] if len(row) > 3 else None
                    cust_id = get_or_create_customer(c_name or f"Token #{token_num}", mob)

                    for col_idx, exact_date in date_cols:
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val and str(val).strip():
                                val_str = str(val).strip()
                                if val_str.replace(".","").isdigit():
                                    amt = float(val_str)
                                    if amt > 0:
                                        rec = f"REC-INSTALL-{token_num}-{exact_date.strftime('%Y%m')}"
                                        installments_batch.append((cust_id, 1, comm_id, exact_date.month, exact_date.year, amt, exact_date, 'cash', rec, exact_date, str(token_num), comm_id))
                                else:
                                    lotteries_batch.append((comm_id, cust_id, exact_date, f"Winner Reward: {val_str}", 'completed', exact_date))

    print(f"\nExecuting fast batch insertion...")

    if collections_batch:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO collections (customer_id, collector_id, branch_id, amount, payment_mode, receipt_number, notes, verification_status, collected_at, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, collections_batch, page_size=1000)
        print(f"  Inserted {len(collections_batch)} collections in batch.")

    if lotteries_batch:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO lotteries (committee_id, winner_id, draw_date, notes, status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, lotteries_batch, page_size=1000)
        print(f"  Inserted {len(lotteries_batch)} lotteries/winners in batch.")

    if installments_batch:
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO installments (customer_id, collector_id, token_id, committee_id, month, year, amount, payment_date, payment_mode, receipt_number, created_at)
            SELECT %s, %s, id, %s, %s, %s, %s, %s, %s, %s, %s
            FROM tokens WHERE token_number = %s AND committee_id = %s LIMIT 1
        """, installments_batch, page_size=1000)
        print(f"  Inserted {len(installments_batch)} installments in batch.")

    conn.commit()
    print("\nFast batch import finished successfully!")
    cur.close()
    conn.close()

if __name__ == "__main__":
    run_fast_batch_import()

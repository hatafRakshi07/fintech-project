#!/usr/bin/env python3
"""
BISSI ERP - Complete Database Rebuild and Import Script
Source of Truth: Bissi folder (1).xlsx
"""

import openpyxl
import psycopg2
import psycopg2.extras
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

# ============================================================
# CONFIG
# ============================================================
WORKBOOK_PATH = 'C:/Users/iSN_kota_T52/Downloads/Bissi folder (1).xlsx'
DB_URL = 'postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require'

COMMITTEE_MAP = {
    'BISSI-1': '11111111-1111-1111-1111-111111111111',  # Hare Ka Sahara (20th Date)
    'BISSI-2': '22222222-2222-2222-2222-222222222222',  # Shree Krishna Associates
    'BISSI-3': '33333333-3333-3333-3333-333333333333',  # Pyare Mohan (15th Date)
    'BISSI-4': 'a3d68b9c-63df-4884-a5ad-eb8a17e3be31',  # Sawariya Seth (5th Date)
}

# Map old integer committee_id to code for migration
OLD_INT_TO_CODE = {1: 'BISSI-1', 2: 'BISSI-2', 3: 'BISSI-3', 4: 'BISSI-4'}

REQUIRED_SHEETS = [
    'Sawariya seth 5 date',
    'Sawariya bissi 5 date gift shee',
    'Sawariya seth bissi gift record',
    'Pyare Mohan bissi gift sheets',
    'Pyare mohan 15 date',
    'Pyare mohan bissi gift records',
    'Hare ka sahara bissi gift sheet',
    'Hare ka sahara bissi maturity a',
    'Hare ka sahara bissi 20 date',
    'Hare ka sahara bissi gift recor',
    'Shree krishna gift sheet',
    'Shree Krishna associate lottery',
    'Shree krishna aasociates gift r',
    'Special customer token no in ea',
    'OUTER Customers list',
    ' monthly payment details',
    'other pending amounts',
    'Lucky Token list',
    'daily diary',
]

# Bissi sheet config: (sheet_name, committee_code, data_start_col_0idx)
BISSI_SHEETS = [
    ('Hare ka sahara bissi 20 date', 'BISSI-1', 7),     # col 7 onwards = dates
    ('Shree Krishna associate lottery', 'BISSI-2', 6),   # col 6 onwards = dates
    ('Pyare mohan 15 date', 'BISSI-3', 9),               # col 9 onwards = dates
    ('Sawariya seth 5 date', 'BISSI-4', 7),              # col 7 onwards = dates
]

# Gift record sheet config: (sheet_name, committee_code, data_start_col_0idx)
GIFT_RECORD_SHEETS = [
    ('Hare ka sahara bissi gift recor', 'BISSI-1', 5),   # col 5 = first date/gift col
    ('Shree krishna aasociates gift r', 'BISSI-2', 5),   # col 5 = first date col
    ('Pyare mohan bissi gift records', 'BISSI-3', 5),    # col 5 = MARCH text
    ('Sawariya seth bissi gift record', 'BISSI-4', 5),   # col 5 = first date col
]


# ============================================================
# HELPERS
# ============================================================

def parse_amount(val):
    """Parse a cell value as a numeric amount, return None if invalid."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ['-', '', 'None', 'done', 'start from', 'Pending', '#VALUE!', '#REF!', '#NAME?', '#N/A']:
        return None
    # Strip currency symbols and whitespace
    s = s.replace('₹', '').replace(',', '').replace(' ', '')
    # Handle partial text like "1000 cash for fan" - extract first number
    import re
    m = re.match(r'^-?[\d]+\.?\d*', s)
    if m:
        try:
            amt = Decimal(m.group())
            return float(amt) if amt > 0 else None
        except InvalidOperation:
            return None
    return None


def parse_date_cell(val):
    """Parse date from cell value, return date or None."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        if isinstance(val, datetime):
            return val.date()
        return val
    s = str(val).strip()
    if not s or s == '-':
        return None
    import re
    # Try DD-MM-YYYY or DD/MM/YYYY
    m = re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})', s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    # Try YYYY-MM-DD
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def normalize_token(val):
    """Parse token number from cell value."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ['-', '', 'None']:
        return None
    import re
    # Handle "1/2" fraction tokens - skip them
    if '/' in s and not s.replace('/', '').replace('.', '').isdigit():
        return None
    # Extract leading integer
    m = re.match(r'^(\d+)', s.replace('.0', ''))
    if m:
        return int(m.group(1))
    return None


def normalize_gift_name(val):
    """Normalize a gift description."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ['-', '', 'None']:
        return None
    return s


# ============================================================
# SCHEMA MIGRATION
# ============================================================

def get_connection():
    """Create a fresh database connection."""
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    return conn


def run_migration(conn):
    """Run schema migration as complete statements."""
    print("\n--- Step 1: Schema Migration ---")
    
    migrations = [
        "ALTER TABLE collections ADD COLUMN IF NOT EXISTS committee_uuid UUID",
        "ALTER TABLE collections ADD COLUMN IF NOT EXISTS customer_uuid UUID",
        "ALTER TABLE collections ADD COLUMN IF NOT EXISTS token_uuid UUID",
        "ALTER TABLE lotteries ADD COLUMN IF NOT EXISTS committee_uuid UUID",
        "ALTER TABLE lotteries ADD COLUMN IF NOT EXISTS winner_token_uuid UUID",
        "ALTER TABLE lotteries ADD COLUMN IF NOT EXISTS winner_customer_uuid UUID",
        "ALTER TABLE lotteries ADD COLUMN IF NOT EXISTS draw_month TEXT",
        "ALTER TABLE lotteries ADD COLUMN IF NOT EXISTS token_number INTEGER",
        "ALTER TABLE lotteries ADD COLUMN IF NOT EXISTS reward_description TEXT",
        "ALTER TABLE gift_distributions ADD COLUMN IF NOT EXISTS committee_uuid UUID",
        "ALTER TABLE gift_distributions ADD COLUMN IF NOT EXISTS customer_uuid UUID",
        "ALTER TABLE gift_distributions ADD COLUMN IF NOT EXISTS token_uuid UUID",
        "ALTER TABLE gift_distributions ADD COLUMN IF NOT EXISTS gift_name TEXT",
        "ALTER TABLE gift_distributions ADD COLUMN IF NOT EXISTS token_number INTEGER",
        "ALTER TABLE gift_distributions ADD COLUMN IF NOT EXISTS customer_name TEXT",
        "ALTER TABLE committees ADD COLUMN IF NOT EXISTS bissi_int_id INTEGER",
        """CREATE TABLE IF NOT EXISTS gift_categories (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""",
    ]
    
    for sql in migrations:
        cur = conn.cursor()
        try:
            cur.execute(sql)
            conn.commit()
        except Exception as e:
            conn.rollback()
    
    cur = conn.cursor()
    cur.execute("UPDATE committees SET bissi_int_id = 1 WHERE id = '11111111-1111-1111-1111-111111111111'")
    cur.execute("UPDATE committees SET bissi_int_id = 2 WHERE id = '22222222-2222-2222-2222-222222222222'")
    cur.execute("UPDATE committees SET bissi_int_id = 3 WHERE id = '33333333-3333-3333-3333-333333333333'")
    cur.execute("UPDATE committees SET bissi_int_id = 4 WHERE id = 'a3d68b9c-63df-4884-a5ad-eb8a17e3be31'")
    conn.commit()
    print("  Schema migration complete")


def clear_data(conn):
    """Clear incorrectly linked data."""
    print("\n--- Step 2: Clearing Incorrect Data ---")
    clear_stmts = [
        "DELETE FROM collections",
        "DELETE FROM lotteries",
        "DELETE FROM gift_distributions",
        "DELETE FROM daily_diary_payments",
        "DELETE FROM daily_diary_loans",
    ]
    cur = conn.cursor()
    for stmt in clear_stmts:
        try:
            cur.execute(stmt)
            conn.commit()
            print(f"  {stmt} - OK")
        except Exception as e:
            conn.rollback()
            print(f"  {stmt} - WARN: {e}")


# ============================================================
# ============================================================
# MAIN IMPORT
# ============================================================

class BissiImporter:
    def __init__(self, conn, wb):
        self.conn = conn
        self.wb = wb
        self.cur = conn.cursor()
        self.stats = {
            'collections': 0,
            'lotteries': 0,
            'gifts': 0,
            'daily_diary': 0,
            'errors': 0,
            'skipped': 0,
        }
        # Token lookup cache: (committee_uuid, token_number) -> (token_uuid, customer_uuid, customer_name)
        self._token_cache = {}
        self._load_token_cache()

    def _load_token_cache(self):
        print("Loading token cache from database...")
        self.cur.execute("""
            SELECT t.id, t.committee_id, t.normalized_token_number, t.customer_id, 
                   c.name as customer_name, c.mobile
            FROM tokens t
            JOIN customers c ON c.id = t.customer_id
            ORDER BY t.committee_id, t.normalized_token_number
        """)
        for row in self.cur.fetchall():
            token_uuid, comm_uuid, tok_num, cust_uuid, cust_name, mobile = row
            key = (str(comm_uuid), tok_num)
            self._token_cache[key] = (str(token_uuid), str(cust_uuid), cust_name or '')
        print(f"  Loaded {len(self._token_cache)} tokens")

    def get_token(self, committee_code, token_number):
        """Get (token_uuid, customer_uuid, customer_name) for a given committee+token."""
        comm_uuid = COMMITTEE_MAP.get(committee_code)
        if not comm_uuid:
            return None
        return self._token_cache.get((comm_uuid, token_number))

    def import_bissi_collections(self):
        """Import payment collections from all 4 Bissi sheets."""
        print("\n=== IMPORTING BISSI COLLECTIONS ===")
        total = 0
        
        for sheet_name, comm_code, data_start_col in BISSI_SHEETS:
            print(f"\n  Sheet: {sheet_name} ({comm_code})")
            if sheet_name not in self.wb.sheetnames:
                print(f"    WARNING: Sheet not found!")
                continue
            
            ws = self.wb[sheet_name]
            comm_uuid = COMMITTEE_MAP[comm_code]
            
            # Read header row to get date columns
            rows_iter = ws.iter_rows(min_row=1, values_only=True)
            header = next(rows_iter)
            
            # Build date column list
            date_cols = []
            for i in range(data_start_col, len(header) if header else 0):
                val = header[i] if i < len(header) else None
                d = parse_date_cell(val)
                if d:
                    date_cols.append((i, d))
            
            print(f"    Found {len(date_cols)} date columns: {[str(d[1]) for d in date_cols[:5]]}...")
            
            sheet_count = 0
            skip_count = 0
            batch = []
            
            for row in rows_iter:
                if not row or row[0] is None:
                    continue
                
                token_num = normalize_token(row[0])
                if token_num is None:
                    skip_count += 1
                    continue
                
                token_info = self.get_token(comm_code, token_num)
                if not token_info:
                    skip_count += 1
                    continue
                
                token_uuid, cust_uuid, cust_name = token_info
                
                for col_idx, collection_date in date_cols:
                    if col_idx >= len(row):
                        continue
                    
                    amount = parse_amount(row[col_idx])
                    if amount is None:
                        continue
                    
                    batch.append({
                        'amount': amount,
                        'date': collection_date,
                        'customer_uuid': cust_uuid,
                        'committee_uuid': comm_uuid,
                        'token_uuid': token_uuid,
                        'token_num': token_num,
                    })
                    sheet_count += 1
                    
                    if len(batch) >= 1000:
                        self._flush_collections_batch(batch)
                        total += len(batch)
                        print(f"    ...{total} payments", flush=True)
                        batch = []
            
            if batch:
                self._flush_collections_batch(batch)
                total += len(batch)
            
            print(f"    Imported {sheet_count} payments, skipped {skip_count} tokens")
        
        self.stats['collections'] = total
        print(f"\n  Total collections imported: {total}")

    def _flush_collections_batch(self, batch):
        """Insert a batch of collection records using fast execute_values."""
        psycopg2.extras.execute_values(
            self.cur,
            """
            INSERT INTO collections 
                (customer_id, committee_id, amount, payment_mode, collected_at, created_at,
                 customer_uuid, committee_uuid, token_uuid, notes, verification_status)
            VALUES %s
            """,
            [
                (
                    1, 1,
                    r['amount'], 'cash', r['date'], datetime.now(),
                    r['customer_uuid'], r['committee_uuid'], r['token_uuid'],
                    f"Token {r['token_num']}", 'verified',
                )
                for r in batch
            ],
            template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            page_size=500
        )
        self.conn.commit()
        sys.stdout.flush()

    def import_lotteries(self):
        """Import lucky draw results from Lucky Token list sheet."""
        print("\n=== IMPORTING LUCKY DRAWS ===")
        
        sheet_name = 'Lucky Token list'
        if sheet_name not in self.wb.sheetnames:
            print("  WARNING: Lucky Token list sheet not found!")
            return
        
        ws = self.wb[sheet_name]
        
        # The Lucky Token list has a specific layout:
        # Row 1: scheme headers (Sawariya Seth vc, Shree krishna associate lotrey, Pyare Mohan Vc, Hare ka Sahara vc)
        # Row 2: Month, Token no columns (repeated 4 times)
        # Row 3+: data rows
        
        # Column mapping:
        # Col 0-1: Sawariya Seth (Month, Token no)
        # Col 3-4: Shree Krishna (Month, Token no)
        # Col 6-7: Pyare Mohan (Month, Token no)
        # Col 9-10: Hare Ka Sahara (Month, Token no)
        
        LUCKY_COMM_MAP = [
            (0, 1, 'BISSI-4'),   # Sawariya Seth
            (3, 4, 'BISSI-2'),   # Shree Krishna
            (6, 7, 'BISSI-3'),   # Pyare Mohan
            (9, 10, 'BISSI-1'),  # Hare Ka Sahara
        ]
        
        rows = list(ws.iter_rows(min_row=3, values_only=True))  # Skip headers
        count = 0
        
        for row in rows:
            for month_col, token_col, comm_code in LUCKY_COMM_MAP:
                if month_col >= len(row) or token_col >= len(row):
                    continue
                
                month_val = row[month_col]
                token_val = row[token_col]
                
                if month_val is None or token_val is None:
                    continue
                
                draw_date = parse_date_cell(month_val)
                if not draw_date:
                    continue
                
                comm_uuid = COMMITTEE_MAP[comm_code]
                
                # Token value can be "260/461" (two tokens) or single like "617"
                token_str = str(token_val).strip()
                token_nums = []
                import re
                for part in re.split(r'[/,\s]+', token_str):
                    t = normalize_token(part)
                    if t:
                        token_nums.append(t)
                
                for tok_num in token_nums:
                    token_info = self.get_token(comm_code, tok_num)
                    if not token_info:
                        continue
                    
                    token_uuid, cust_uuid, cust_name = token_info
                    
                    try:
                        self.cur.execute("""
                            INSERT INTO lotteries 
                                (committee_id, draw_date, winner_id, status, notes,
                                 committee_uuid, winner_token_uuid, winner_customer_uuid,
                                 draw_month, token_number, reward_description)
                            VALUES (
                                1, %s, 1, 'completed', %s,
                                %s::uuid, %s::uuid, %s::uuid,
                                %s, %s, 'Lucky'
                            )
                        """, (
                            draw_date,
                            f'Lucky Draw - {comm_code} - Token {tok_num} - {cust_name}',
                            comm_uuid, token_uuid, cust_uuid,
                            draw_date.strftime('%Y-%m'), tok_num
                        ))
                        count += 1
                    except Exception as e:
                        self.stats['errors'] += 1
        
        self.conn.commit()
        self.stats['lotteries'] = count
        print(f"  Imported {count} lucky draw records")

    def import_gift_records(self):
        """Import gift distributions from gift record sheets."""
        print("\n=== IMPORTING GIFT RECORDS ===")
        total = 0
        
        for sheet_name, comm_code, data_start_col in GIFT_RECORD_SHEETS:
            print(f"\n  Sheet: {sheet_name} ({comm_code})")
            if sheet_name not in self.wb.sheetnames:
                print(f"    WARNING: Sheet not found!")
                continue
            
            ws = self.wb[sheet_name]
            comm_uuid = COMMITTEE_MAP[comm_code]
            
            # Read header row to get date columns
            rows_iter = ws.iter_rows(min_row=1, values_only=True)
            header = next(rows_iter)
            
            # Build date/period column list
            gift_cols = []
            for i in range(data_start_col, len(header) if header else 0):
                val = header[i] if i < len(header) else None
                if val is None:
                    continue
                d = parse_date_cell(val)
                if d:
                    gift_cols.append((i, d, str(d)))
                elif str(val).strip().upper() not in ['-', '', 'NONE']:
                    # Text like "MARCH" - use as period label
                    gift_cols.append((i, None, str(val).strip()))
            
            print(f"    Gift columns: {len(gift_cols)}")
            
            sheet_count = 0
            
            for row in rows_iter:
                if not row or row[0] is None:
                    continue
                
                token_num = normalize_token(row[0])
                if token_num is None:
                    continue
                
                token_info = self.get_token(comm_code, token_num)
                if not token_info:
                    continue
                
                token_uuid, cust_uuid, cust_name = token_info
                
                for col_idx, col_date, col_label in gift_cols:
                    if col_idx >= len(row):
                        continue
                    
                    gift_val = normalize_gift_name(row[col_idx])
                    if not gift_val:
                        continue
                    
                    # Use col_date or derive from label
                    dist_date = col_date or date.today()
                    
                    try:
                        self.cur.execute("""
                            INSERT INTO gift_distributions 
                                (customer_id, committee_id, gift_id, distribution_date, status,
                                 notes, branch_id,
                                 committee_uuid, customer_uuid, token_uuid,
                                 gift_name, token_number, customer_name)
                            VALUES (
                                1, 1, 1, %s, 'DELIVERED',
                                %s, 1,
                                %s::uuid, %s::uuid, %s::uuid,
                                %s, %s, %s
                            )
                        """, (
                            dist_date,
                            f'Period: {col_label}',
                            comm_uuid, cust_uuid, token_uuid,
                            gift_val, token_num, cust_name
                        ))
                        sheet_count += 1
                    except Exception as e:
                        self.stats['errors'] += 1
            
            self.conn.commit()
            print(f"    Imported {sheet_count} gift records")
            total += sheet_count
        
        # Also import from gift distribution sheets (grouped by month)
        total += self.import_gift_distribution_sheets()
        
        self.stats['gifts'] = total
        print(f"\n  Total gift records imported: {total}")

    def import_gift_distribution_sheets(self):
        """Import from the grouped gift distribution sheets (Lucky + gift columns)."""
        GIFT_DIST_SHEETS = [
            ('Sawariya bissi 5 date gift shee', 'BISSI-4'),
            ('Pyare Mohan bissi gift sheets', 'BISSI-3'),
            ('Hare ka sahara bissi gift sheet', 'BISSI-1'),
            ('Shree krishna gift sheet', 'BISSI-2'),
        ]
        
        total = 0
        for sheet_name, comm_code in GIFT_DIST_SHEETS:
            if sheet_name not in self.wb.sheetnames:
                continue
            
            print(f"\n  Gift dist sheet: {sheet_name}")
            ws = self.wb[sheet_name]
            comm_uuid = COMMITTEE_MAP[comm_code]
            
            # These sheets have format:
            # Row 1: [Date, -, -, -, Date, -, -, -, ...]   <- draw dates
            # Row 2: [Name, Token No, Gift Status, -, ...]  <- column headers (repeated)
            # Row 3+: [name, token, gift_name, -, ...]     <- data
            
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if len(rows) < 3:
                continue
            
            header_row1 = rows[0]  # Draw dates
            header_row2 = rows[1]  # Name/Token/Gift headers
            
            # Find the blocks: each block is 4 columns wide (Name, Token No, Gift Status, -)
            # Find date positions in row 1
            block_starts = []
            for i, val in enumerate(header_row1):
                d = parse_date_cell(val)
                if d:
                    block_starts.append((i, d))
                elif val and str(val).strip() not in ['-', '', 'None']:
                    # Could be a text date like "Septmber - 24"
                    import re
                    m = re.search(r'(\w+)\s*-?\s*(\d{2,4})', str(val).strip())
                    if m:
                        block_starts.append((i, None))
            
            if not block_starts:
                continue
            
            sheet_count = 0
            for data_row in rows[2:]:
                if not data_row:
                    continue
                
                for b_idx, (col_start, draw_date) in enumerate(block_starts):
                    # Each block: col_start = Name, col_start+1 = Token No, col_start+2 = Gift Status
                    name_col = col_start
                    token_col = col_start + 1
                    gift_col = col_start + 2
                    
                    if token_col >= len(data_row) or gift_col >= len(data_row):
                        continue
                    
                    token_val = data_row[token_col]
                    gift_val = normalize_gift_name(data_row[gift_col])
                    
                    if gift_val is None:
                        continue
                    
                    token_num = normalize_token(token_val)
                    if not token_num:
                        continue
                    
                    token_info = self.get_token(comm_code, token_num)
                    if not token_info:
                        continue
                    
                    token_uuid, cust_uuid, cust_name = token_info
                    dist_date = draw_date or date.today()
                    
                    # Mark as Lucky if gift_val == 'Lucky'
                    is_lucky = gift_val.lower() in ['lucky', 'lucky ']
                    
                    try:
                        self.cur.execute("""
                            INSERT INTO gift_distributions 
                                (customer_id, committee_id, gift_id, distribution_date, status,
                                 notes, branch_id,
                                 committee_uuid, customer_uuid, token_uuid,
                                 gift_name, token_number, customer_name)
                            VALUES (
                                1, 1, 1, %s, %s,
                                %s, 1,
                                %s::uuid, %s::uuid, %s::uuid,
                                %s, %s, %s
                            )
                        """, (
                            dist_date,
                            'PENDING' if is_lucky else 'DELIVERED',
                            f'Gift dist: {draw_date}',
                            comm_uuid, cust_uuid, token_uuid,
                            gift_val, token_num, cust_name
                        ))
                        sheet_count += 1
                    except Exception as e:
                        self.stats['errors'] += 1
            
            self.conn.commit()
            print(f"    {sheet_count} gift dist records from {sheet_name}")
            total += sheet_count
        
        return total

    def import_daily_diary(self):
        """Import daily diary loans from the daily diary sheet."""
        print("\n=== IMPORTING DAILY DIARY ===")
        
        sheet_name = 'daily diary'
        if sheet_name not in self.wb.sheetnames:
            print("  WARNING: daily diary sheet not found!")
            return
        
        ws = self.wb[sheet_name]
        count = 0
        
        # daily diary columns (from analysis):
        # 0=NAME, 1=MOBILE NO, 2=REF MOBILE, 3=Payment Mode, 4=REASON, 5=ADDRESS,
        # 6=SECURITY, 7=Loan AMOUNT, 8=START DATE, 9=COMPLETE DATE,
        # 10=AMOUNT TAKEN, 11=REMAINING TILL TODAY, 12=TOTAL REMAINING
        
        rows_iter = ws.iter_rows(min_row=2, values_only=True)  # Skip header
        
        for row in rows_iter:
            if not row or row[0] is None:
                continue
            
            name_val = str(row[0]).strip() if row[0] else None
            if not name_val or name_val == '-':
                continue
            
            mobile = str(row[1]).strip() if row[1] else None
            if mobile:
                import re
                mobile = re.sub(r'[^\d]', '', str(mobile))[:10]
            
            ref_mobile = str(row[2]).strip() if row[2] else None
            address = str(row[5]).strip() if row[5] else None
            security = str(row[6]).strip() if row[6] else None
            
            loan_amount = parse_amount(row[7])
            if not loan_amount:
                continue
            
            start_date = parse_date_cell(row[8])
            complete_date_raw = row[9]
            
            complete_date = parse_date_cell(complete_date_raw)
            if not complete_date and complete_date_raw:
                # Could be text like "BEFORE 29/12/26"
                import re
                m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', str(complete_date_raw))
                if m:
                    try:
                        yr = int(m.group(3))
                        if yr < 100:
                            yr += 2000
                        complete_date = date(yr, int(m.group(2)), int(m.group(1)))
                    except:
                        pass
            
            amount_taken = parse_amount(row[10])
            remaining_till_today = parse_amount(row[11]) if row[11] else None
            total_remaining = parse_amount(row[12]) if row[12] else None
            
            notes = str(row[4]).strip() if row[4] else None
            
            # Extract collection plan from name (e.g., "(5)" means 500/day)
            collection_plan = '500/day'  # default
            import re
            plan_match = re.search(r'\((\d+)\)', name_val)
            if plan_match:
                plan_num = int(plan_match.group(1))
                # Numbers in name like (5), (6), (8) typically mean plan amount × 100
                common_plans = {5: '500/day', 6: '600/day', 8: '800/day', 
                               10: '1000/day', 15: '1500/day', 20: '2000/day'}
                collection_plan = common_plans.get(plan_num, f'{plan_num}00/day')
            
            # Compute status
            if total_remaining is not None and total_remaining <= 0:
                status = 'COMPLETED'
            elif remaining_till_today is not None:
                status = 'ACTIVE'
            else:
                status = 'ACTIVE'
            
            try:
                self.cur.execute("""
                    INSERT INTO daily_diary_loans 
                        (customer_name, mobile_number, reference_mobile_numbers,
                         address, security, loan_amount, start_date, 
                         expected_complete_date, collection_plan, notes, status,
                         created_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    RETURNING id
                """, (
                    name_val, mobile, ref_mobile,
                    address, security, loan_amount, start_date,
                    complete_date, collection_plan, notes, status
                ))
                loan_id = self.cur.fetchone()[0]
                
                # If there's amount_taken, add a payment record
                if amount_taken and amount_taken > 0 and start_date:
                    self.cur.execute("""
                        INSERT INTO daily_diary_payments
                            (loan_id, amount_deposited, payment_date, payment_mode,
                             running_remaining_balance, created_at)
                        VALUES (%s, %s, %s, 'Cash', %s, NOW())
                    """, (
                        loan_id, amount_taken, start_date,
                        total_remaining or (loan_amount - amount_taken)
                    ))
                
                count += 1
            except Exception as e:
                print(f"    ERROR on daily diary row: {name_val} - {e}")
                self.stats['errors'] += 1
        
        self.conn.commit()
        self.stats['daily_diary'] = count
        print(f"  Imported {count} daily diary loans")

    def print_summary(self):
        print("\n" + "="*60)
        print("IMPORT SUMMARY")
        print("="*60)
        print(f"  Collections:    {self.stats['collections']}")
        print(f"  Lotteries:      {self.stats['lotteries']}")
        print(f"  Gift Records:   {self.stats['gifts']}")
        print(f"  Daily Diary:    {self.stats['daily_diary']}")
        print(f"  Errors:         {self.stats['errors']}")
        print(f"  Skipped:        {self.stats['skipped']}")
        print("="*60)


def main():
    print("=" * 60)
    print("BISSI ERP - COMPLETE DATABASE REBUILD & IMPORT")
    print("=" * 60)
    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Database: Neon PostgreSQL")

    # Load workbook
    print("\nLoading workbook...")
    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
        print(f"Loaded {len(wb.sheetnames)} worksheets")
        missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            print(f"WARNING: Missing sheets: {missing}")
    except Exception as e:
        print(f"ERROR loading workbook: {e}")
        sys.exit(1)

    conn = None
    try:
        # Connect
        print("\nConnecting to database...")
        conn = get_connection()
        print("Connected successfully")

        # Step 1: Schema migration
        run_migration(conn)

        # Step 2: Clear incorrect data
        clear_data(conn)

        # Reconnect fresh after clearing
        conn.close()
        conn = get_connection()

        # Step 3-7: Import data
        importer = BissiImporter(conn, wb)
        importer.import_bissi_collections()
        importer.import_lotteries()
        importer.import_gift_records()
        importer.import_daily_diary()
        importer.print_summary()

    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            try:
                conn.rollback()
            except:
                pass
        sys.exit(1)
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass
        wb.close()

    print("\n Import completed successfully!")


if __name__ == '__main__':
    main()

"""Import only gift records and daily diary from the Bissi Excel workbook."""
import openpyxl
import psycopg2
import psycopg2.extras
import sys
from datetime import date, datetime
import re

WORKBOOK_PATH = 'C:/Users/iSN_kota_T52/Downloads/Bissi folder (1).xlsx'
DB_URL = 'postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require'

COMMITTEE_MAP = {
    'BISSI-1': '11111111-1111-1111-1111-111111111111',
    'BISSI-2': '22222222-2222-2222-2222-222222222222',
    'BISSI-3': '33333333-3333-3333-3333-333333333333',
    'BISSI-4': 'a3d68b9c-63df-4884-a5ad-eb8a17e3be31',
}

GIFT_RECORD_SHEETS = [
    ('Hare ka sahara bissi gift recor', 'BISSI-1', 5),
    ('Shree krishna aasociates gift r', 'BISSI-2', 5),
    ('Pyare mohan bissi gift records', 'BISSI-3', 5),
    ('Sawariya seth bissi gift record', 'BISSI-4', 5),
]

def normalize_token(val):
    if val is None: return None
    s = str(val).strip()
    if s in ['-', '', 'None']: return None
    m = re.match(r'^(\d+)', s.replace('.0', ''))
    return int(m.group(1)) if m else None

def normalize_gift(val):
    if val is None: return None
    s = str(val).strip()
    return None if s in ['-', '', 'None'] else s

def parse_date_cell(val):
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    return None

def main():
    print('Loading workbook (full load)...')
    # Use data_only=True but NOT read_only - eager load for faster processing
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    
    # PRE-LOAD ALL GIFT DATA INTO MEMORY before connecting to DB
    print('Pre-processing gift records from workbook...')
    gift_rows = []  # list of (comm_uuid, tok_num, dist_date, gift_val)
    
    for sheet_name, comm_code, data_start_col in GIFT_RECORD_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f'Sheet not found: {sheet_name}')
            continue
        
        ws = wb[sheet_name]
        comm_uuid = COMMITTEE_MAP[comm_code]
        header = [ws.cell(1, i+1).value for i in range(ws.max_column)]
        
        gift_cols = []
        for i in range(data_start_col, len(header)):
            val = header[i]
            if val is None: continue
            d = parse_date_cell(val)
            if d:
                gift_cols.append((i, d))
            elif str(val).strip().upper() not in ['-', '', 'NONE']:
                gift_cols.append((i, None))
        
        print(f'  {sheet_name}: {ws.max_row-1} tokens, {len(gift_cols)} gift cols')
        
        for row_idx in range(2, ws.max_row + 1):
            tok_val = ws.cell(row_idx, 1).value
            tok_num = normalize_token(tok_val)
            if not tok_num: continue
            
            for col_idx, col_date in gift_cols:
                cell_val = ws.cell(row_idx, col_idx + 1).value
                gift_val = normalize_gift(cell_val)
                if not gift_val: continue
                
                dist_date = col_date or date.today()
                gift_rows.append((comm_uuid, tok_num, dist_date, gift_val))
    
    print(f'Found {len(gift_rows)} gift records to import')
    
    # PRE-LOAD DAILY DIARY DATA
    print('Pre-processing daily diary...')
    diary_rows = []
    ws_diary = wb['daily diary']
    for row_idx in range(2, ws_diary.max_row + 1):
        name_val = ws_diary.cell(row_idx, 1).value
        if not name_val or str(name_val).strip() in ['-', '', 'None']: continue
        name_str = str(name_val).strip()
        
        def get(col): return ws_diary.cell(row_idx, col).value
        
        mobile = None
        if get(2): mobile = re.sub(r'[^\d]', '', str(get(2)))[:10]
        
        loan_amt = None
        if get(8):
            try: loan_amt = float(str(get(8)).replace(',','').strip())
            except: pass
        if not loan_amt: continue
        
        start_date = None
        v = get(9)
        if v and isinstance(v, (date, datetime)):
            start_date = v.date() if isinstance(v, datetime) else v
        
        complete_date = None
        v2 = get(10)
        if v2:
            if isinstance(v2, (date, datetime)):
                complete_date = v2.date() if isinstance(v2, datetime) else v2
            else:
                m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', str(v2))
                if m:
                    yr = int(m.group(3))
                    if yr < 100: yr += 2000
                    try: complete_date = date(yr, int(m.group(2)), int(m.group(1)))
                    except: pass
        
        amount_taken = None
        if get(11):
            try: amount_taken = float(str(get(11)).replace(',','').strip())
            except: pass
        
        total_remaining = None
        if get(13):
            try: total_remaining = float(str(get(13)).replace(',','').strip())
            except: pass
        
        plan_match = re.search(r'\((\d+)\)', name_str)
        plan = '500/day'
        if plan_match:
            n = int(plan_match.group(1))
            plans = {5:'500/day', 6:'600/day', 8:'800/day', 10:'1000/day', 15:'1500/day', 20:'2000/day'}
            plan = plans.get(n, f'{n}00/day')
        
        status = 'COMPLETED' if total_remaining is not None and total_remaining <= 0 else 'ACTIVE'
        notes = str(get(5)).strip() if get(5) else None
        address = str(get(6)).strip() if get(6) else None
        
        diary_rows.append((name_str, mobile, str(get(3)).strip() if get(3) else None,
                           address, None, loan_amt, start_date, complete_date,
                           plan, notes, status, amount_taken, total_remaining))
    
    print(f'Found {len(diary_rows)} diary loans')
    wb.close()
    
    # NOW CONNECT AND DO DB OPERATIONS
    print('Connecting...')
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    
    # Load token cache
    print('Loading token cache...')
    cur.execute('SELECT t.id, t.committee_id, t.normalized_token_number, t.customer_id, c.name FROM tokens t JOIN customers c ON c.id = t.customer_id')
    token_cache = {}
    for tok_id, comm_id, tok_num, cust_id, cust_name in cur.fetchall():
        key = (str(comm_id), tok_num)
        token_cache[key] = (str(tok_id), str(cust_id), cust_name or '')
    print(f'  Loaded {len(token_cache)} tokens')
    
    # Clear existing
    cur.execute('DELETE FROM gift_distributions')
    cur.execute('DELETE FROM daily_diary_loans')
    cur.execute('DELETE FROM daily_diary_payments')
    conn.commit()
    
    # Batch insert gifts
    print('Inserting gift records...')
    gift_batch = []
    for comm_uuid, tok_num, dist_date, gift_val in gift_rows:
        info = token_cache.get((comm_uuid, tok_num))
        if not info: continue
        tok_uuid, cust_uuid, cust_name = info
        gift_batch.append((1, 1, 1, dist_date, 'given', f'Bissi gift', 1, 
                          comm_uuid, cust_uuid, tok_uuid, gift_val, tok_num, cust_name))
    
    if gift_batch:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO gift_distributions
               (customer_id, committee_id, gift_id, distribution_date, status, notes, branch_id,
                committee_uuid, customer_uuid, token_uuid, gift_name, token_number, customer_name)
               VALUES %s""",
            gift_batch,
            template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            page_size=200
        )
        conn.commit()
        print(f'Inserted {len(gift_batch)} gift records')
    
    # Insert diary
    print('Inserting diary loans...')
    diary_count = 0
    for row in diary_rows:
        name, mobile, ref_mobile, address, security, loan_amt, start_date, complete_date, plan, notes, status, amount_taken, total_remaining = row
        try:
            cur.execute("""
                INSERT INTO daily_diary_loans
                    (customer_name, mobile_number, reference_mobile_numbers, address, security,
                     loan_amount, start_date, expected_complete_date, collection_plan, notes, status, created_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()) RETURNING id
            """, (name, mobile, ref_mobile, address, security, loan_amt, start_date, complete_date, plan, notes, status))
            loan_id = cur.fetchone()[0]
            if amount_taken and amount_taken > 0 and start_date:
                cur.execute("""
                    INSERT INTO daily_diary_payments
                        (loan_id, amount_deposited, payment_date, payment_mode, notes, created_at)
                    VALUES (%s, %s, %s, 'Cash', %s, NOW())
                """, (loan_id, amount_taken, start_date, f'Initial deposit. Remaining: {total_remaining}'))
            diary_count += 1
        except Exception as e:
            conn.rollback()
            print(f'  Diary error: {name} - {e}')
    
    conn.commit()
    print(f'Inserted {diary_count} diary loans')
    print(f'\n=== DONE: {len(gift_batch)} gifts, {diary_count} diary ===')
    conn.close()

if __name__ == '__main__':
    main()

import os
import re
import sys
import openpyxl
import psycopg2
from datetime import datetime, date

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

NEON_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require"
)
EXCEL_PATH = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder (1).xlsx"

def clean_phone(val):
    if not val: return None
    s = re.sub(r"[^\d]", "", str(val).split(".")[0])
    s = s[-10:] if len(s) >= 10 else s
    return s if len(s) >= 6 else None

def clean_name(val):
    if not val: return None
    v = str(val).strip()
    v = re.sub(r"\s+", " ", v)
    return v if v and v.lower() not in ("none", "name", "name ", "", "jsk", "-") else None

def parse_date(val, default_date=None):
    if not val: return default_date or datetime.now()
    if isinstance(val, datetime): return val
    if isinstance(val, date): return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    parts = re.split(r"[/-]", s)
    if len(parts) == 3:
        try:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100: y += 2000
            return datetime(y, m, d)
        except Exception: pass
    return default_date or datetime.now()

def run_update():
    print("=== UPDATING HARE KA SAHARA BISSI (₹2500) & GIFTS WITH ACCURATE DATES ===")
    conn = psycopg2.connect(NEON_URL)
    cur = conn.cursor()

    # 1. Update Hare Ka Sahara Bissi Installment Amount to ₹2,500 in DB
    print("\nUpdating Hare Ka Sahara Bissi installment amount to ₹2,500...")
    cur.execute("""
        UPDATE committees 
        SET installment_amount = 2500.0, updated_at = NOW() 
        WHERE id = 3 OR name LIKE '%Hare Ka Sahara%'
    """)
    print(f"Updated {cur.rowcount} committee rows.")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    cur.execute("SELECT id, name, mobile FROM customers")
    customers = {}
    for cid, name, mob in cur.fetchall():
        if mob: customers[mob] = cid
        if name: customers[name.lower().strip()] = cid

    def get_or_create_customer(name, mobile=None):
        name_clean = clean_name(name)
        mob_clean = clean_phone(mobile)
        if mob_clean and mob_clean in customers: return customers[mob_clean]
        if name_clean and name_clean.lower() in customers: return customers[name_clean.lower()]

        final_name = name_clean or f"Member #{mob_clean or 'Unknown'}"
        final_mob = mob_clean or f"9000{len(customers)+100000:06d}"
        ref_no = f"REF-{len(customers) + 1001}"
        cur.execute("""
            INSERT INTO customers (name, mobile, reference_number, branch_id, status, created_at, updated_at)
            VALUES (%s, %s, %s, 1, 'active', NOW(), NOW())
            RETURNING id
        """, (final_name, final_mob, ref_no))
        cid = cur.fetchone()[0]
        customers[final_mob] = cid
        if name_clean: customers[name_clean.lower()] = cid
        return cid

    # Ensure Gift Category
    cur.execute("SELECT id FROM gift_categories LIMIT 1")
    cat_row = cur.fetchone()
    if cat_row:
        gift_cat_id = cat_row[0]
    else:
        cur.execute("INSERT INTO gift_categories (name, branch_id) VALUES ('Bissi Rewards', 1) RETURNING id")
        gift_cat_id = cur.fetchone()[0]

    gifts_count = 0
    lotteries_count = 0

    # 2. Process Gift Sheets with exact dates
    gift_sheets_config = [
        ("Shree krishna aasociates gift r", 4, True),  # Shree Krishna Bissi: Only Gifts
        ("Radhe krishna bissi gift list", 4, False),   # Cash & Gifts
        ("Shree Krishna associate lottery", 4, False),
    ]

    for sheet_name, comm_id, gifts_only in gift_sheets_config:
        if sheet_name in wb.sheetnames:
            print(f"\nProcessing gifts from sheet: '{sheet_name}'...")
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                header = rows[0]
                date_cols = []
                for idx, col in enumerate(header):
                    if idx >= 1 and col:
                        dt = parse_date(col, None)
                        if dt: date_cols.append((idx, dt))
                
                for row in rows[1:]:
                    if not row or not row[0]: continue
                    tok_no = str(row[0]).split(".")[0].strip()
                    if not tok_no.isdigit(): continue
                    token_num = int(tok_no)
                    
                    c_name = row[1] if len(row) > 1 else None
                    mob = row[3] if len(row) > 3 else None
                    cust_id = get_or_create_customer(c_name or f"Token #{token_num}", mob)

                    for col_idx, exact_date in date_cols:
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val and str(val).strip():
                                val_str = str(val).strip()
                                # Record Gift / Winner reward
                                cur.execute("""
                                    INSERT INTO lotteries (committee_id, winner_id, draw_date, notes, status, created_at)
                                    VALUES (%s, %s, %s, %s, 'completed', %s)
                                """, (comm_id, cust_id, exact_date, f"Winner Reward: {val_str}", exact_date))
                                lotteries_count += 1

    conn.commit()
    print(f"\nGift & Bissi updates completed successfully!")
    print(f"Total Winners/Lotteries updated: {lotteries_count}")
    print(f"Total Gifts recorded: {gifts_count}")
    cur.close()
    conn.close()

if __name__ == "__main__":
    run_update()

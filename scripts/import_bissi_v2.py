#!/usr/bin/env python3
"""
Import data from Bissi.xlsx into the live Neon database.
- Daily Diary: 18 loan records + payment totals
- BYAJ KI LIST: 298 interest accounts + received payment transactions
"""
import re, sys
from datetime import date, datetime
import openpyxl, psycopg2

DB = "postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed-pooler.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require"
WB = "Bissi.xlsx"

def num(v, d=0.0):
    if v is None: return d
    try: return float(str(v).replace(',',''))
    except: return d

def txt(v):
    if v is None: return ""
    s = str(v).strip()
    return re.sub(r'\.0$', '', s) if re.match(r'^\d+\.0$', s) else s

def to_date(v):
    if v is None: return None
    if isinstance(v, (datetime,)): return v.date()
    if isinstance(v, date): return v
    for fmt in ("%Y-%m-%d","%d/%m/%y","%d/%m/%Y","%m/%d/%Y"):
        try: return datetime.strptime(str(v).strip(), fmt).date()
        except: pass
    return None

def received(v):
    if not v: return False
    return any(k in str(v).lower() for k in ('received','done','rcvd'))

def serial(name):
    m = re.search(r'\((\d+)\)', str(name))
    return int(m.group(1)) if m else None

def main():
    print("Loading", WB)
    wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)

    conn = psycopg2.connect(DB)
    conn.autocommit = False
    cur = conn.cursor()

    # ── 1. Daily Diary ────────────────────────────────────────────────────────
    print("\n[1] Daily Diary loans…")
    cur.execute("DELETE FROM daily_diary_payments")
    cur.execute("DELETE FROM daily_diary_loans")

    ws = wb['daily diary']
    n_loans = n_pay = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or not str(name).strip(): continue
        loan_amt  = num(row[7])
        if loan_amt <= 0: continue

        mobile    = txt(row[1])
        ref_mob   = txt(row[2])
        pay_mode  = txt(row[3])
        reason    = txt(row[4])
        address   = txt(row[5])
        security  = txt(row[6])
        start_dt  = to_date(row[8])
        end_dt    = to_date(row[9])
        taken     = num(row[10])

        remaining = max(0, loan_amt - taken)
        status    = 'COMPLETED' if remaining <= 0 else 'ACTIVE'

        # derive daily collection amount from duration
        plan = "500/day"
        if start_dt and end_dt:
            days = max(1, (end_dt - start_dt).days)
            daily = round(loan_amt / days)
            if daily > 0:
                plan = f"{int(daily)}/day"

        sn = serial(str(name))
        notes = "|".join(filter(None, [f"SN:{sn}" if sn else None, pay_mode, reason]))

        cur.execute("""
            INSERT INTO daily_diary_loans
                (customer_name, mobile_number, reference_mobile_numbers,
                 address, security, loan_amount, start_date,
                 expected_complete_date, collection_plan, status, notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            txt(name), mobile or None, ref_mob or None,
            address or None, security or None,
            loan_amt,
            str(start_dt) if start_dt else None,
            str(end_dt)   if end_dt   else None,
            plan, status, notes or None,
        ))
        lid = cur.fetchone()[0]
        n_loans += 1

        if taken > 0:
            pay_date = end_dt or start_dt or date.today()
            cur.execute("""
                INSERT INTO daily_diary_payments
                    (loan_id, payment_date, amount_deposited, payment_mode, notes)
                VALUES (%s,%s,%s,%s,%s)
            """, (lid, str(pay_date), taken, pay_mode or 'Cash', 'Excel import – total taken'))
            n_pay += 1

    conn.commit()
    print(f"  ✓ {n_loans} loans, {n_pay} payments")

    # ── 2. BYAJ KI LIST ───────────────────────────────────────────────────────
    print("\n[2] BYAJ interest accounts…")
    cur.execute("DELETE FROM interest_transactions")
    cur.execute("DELETE FROM interest_accounts")

    ws2 = wb['BYAJ KI LIST']
    hdr = list(ws2.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Collect date columns (index ≥ 10 that are actual dates)
    date_cols = []
    for i, h in enumerate(hdr):
        if i >= 10:
            d = to_date(h)
            if d: date_cols.append((i, d))

    n_acct = n_txn = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or not str(name).strip(): continue
        monthly_amt = num(row[5])
        if monthly_amt <= 0: continue

        address  = txt(row[1]) or txt(row[9])
        mobile   = txt(row[2])
        ref_mob  = txt(row[3])
        int_date = txt(row[4])
        reply    = txt(row[6])
        reason1  = txt(row[7])

        sn = serial(str(name))
        cust_ph = sn if sn else (n_acct + 1)   # integer placeholder (no FK)
        notes_s = f"{txt(name)} | Mob:{mobile} | Ref:{ref_mob} | Date:{int_date} | {address}"

        # find earliest date in date_cols as start_date
        start_dt = date_cols[0][1] if date_cols else date(2024, 10, 1)

        cur.execute("""
            INSERT INTO interest_accounts
                (customer_id, principal_amount, interest_rate,
                 start_date, total_interest_paid, pending_interest,
                 monthly_interest, outstanding_amount,
                 status, branch_id, notes)
            VALUES (%s, 0, 0, %s, 0, %s, %s, %s, 'active', 1, %s)
            RETURNING id
        """, (
            cust_ph, str(start_dt),
            monthly_amt, monthly_amt, monthly_amt,
            notes_s,
        ))
        acct_id = cur.fetchone()[0]
        n_acct += 1

        # insert received payment transactions
        for col_idx, pay_date in date_cols:
            cell = row[col_idx] if col_idx < len(row) else None
            if not received(cell): continue

            amt_s = re.sub(r'[^0-9.]', '', str(cell))
            try:   amt = float(amt_s) if amt_s else monthly_amt
            except: amt = monthly_amt
            if amt <= 0: amt = monthly_amt

            cur.execute("""
                INSERT INTO interest_transactions
                    (account_id, customer_id, type, amount,
                     month, year, payment_date, branch_id, notes)
                VALUES (%s,%s,'credit',%s,%s,%s,%s,1,%s)
            """, (acct_id, cust_ph, amt,
                  pay_date.month, pay_date.year,
                  str(pay_date), txt(cell)))
            n_txn += 1

    conn.commit()
    print(f"  ✓ {n_acct} interest accounts, {n_txn} transactions")

    # ── 3. Verify ─────────────────────────────────────────────────────────────
    print("\n[3] Final counts:")
    for t in ('daily_diary_loans','daily_diary_payments',
              'interest_accounts','interest_transactions'):
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        print(f"  {t}: {cur.fetchone()[0]}")

    cur.close(); conn.close()
    print("\n✅ Done.")

if __name__ == "__main__":
    main()

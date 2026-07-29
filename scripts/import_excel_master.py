import os
import re
import sys
import openpyxl
import psycopg2
import psycopg2.extras
from datetime import datetime, date

# Enforce UTF-8 output
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

NEON_URL = "postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require"
EXCEL_PATH = r"C:\Users\iSN_kota_T52\Downloads\Bissi folder.xlsx"

def clean_phone(val):
    if not val:
        return None
    s = re.sub(r"[^\d]", "", str(val).split(".")[0])
    s = s[-10:] if len(s) >= 10 else s
    return s if len(s) >= 6 else None

def clean_name(val):
    if not val:
        return None
    v = str(val).strip()
    v = re.sub(r"\s+", " ", v)
    return v if v and v.lower() not in ("none", "name", "name ", "", "jsk", "-") else None

def parse_date(val, default_date=None):
    if not val:
        return default_date or datetime.now()
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    
    s = str(val).strip()
    parts = re.split(r"[/-]", s)
    if len(parts) == 3:
        try:
            d = int(parts[0])
            m = int(parts[1])
            y = int(parts[2])
            if y < 100:
                y += 2000
            return datetime(y, m, d)
        except Exception:
            pass
    return default_date or datetime.now()

def run_master_import():
    print("=== STARTING 7-QUERY SUPERCHARGED BISSI IMPORT ===")
    
    conn = psycopg2.connect(NEON_URL)
    cur = conn.cursor()
    
    # 0. Ensure Branch
    cur.execute("""
        INSERT INTO branches (id, name, code, city, address, status, updated_at)
        VALUES (1, 'Shree Krishna Associate', 'SKA001', 'Main City', 'Main Branch', 'active', NOW())
        ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name
    """)

    # Ensure 4 Committees
    committees_def = [
        (1, 'Sawariya Seth Bissi', 'monthly', 3000.0, 500, 1),
        (2, 'Pyare Mohan Bissi', 'monthly', 3000.0, 500, 1),
        (3, 'Hare Ka Sahara Bissi', 'monthly', 2500.0, 500, 1),
        (4, 'Shree Krishna Bissi', 'monthly', 3000.0, 1111, 1),
    ]
    for c in committees_def:
        cur.execute("""
            INSERT INTO committees (id, name, type, installment_amount, member_limit, branch_id, status, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, 'active', NOW())
            ON CONFLICT (id) DO UPDATE 
            SET name = EXCLUDED.name, installment_amount = EXCLUDED.installment_amount, member_limit = EXCLUDED.member_limit, updated_at = NOW()
        """, c)
    conn.commit()

    print("Loading workbook into memory...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    print("Workbook loaded!")

    # Query 1: Pre-load existing customers in memory
    cur.execute("SELECT id, name, mobile, reference_number FROM customers")
    existing_cust = {}
    for cid, name, mob, ref in cur.fetchall():
        if mob: existing_cust[mob] = cid
        if ref: existing_cust[ref] = cid
        if name: existing_cust[name.lower().strip()] = cid

    # Query 2: Pre-load existing tokens in memory
    cur.execute("SELECT token_number, committee_id, id FROM tokens")
    existing_tokens = {(str(r[0]), r[1]): r[2] for r in cur.fetchall()}

    # Query 3: Pre-load existing memberships in memory
    cur.execute("SELECT committee_id, customer_id, token_number FROM committee_members")
    existing_memberships = {(r[0], r[1], str(r[2])) for r in cur.fetchall()}

    new_customers_dict = {} # ref_num -> (ref_num, name, mob, ref_name, address)

    scheme_sheets = [
        {'name': 'Sawariya seth 5 date', 'committeeId': 1, 'amount': 3000.0},
        {'name': 'Pyare mohan 15 date', 'committeeId': 2, 'amount': 3000.0},
        {'name': 'Hare ka sahara bissi 20 date', 'committeeId': 3, 'amount': 2500.0},
        {'name': 'Shree Krishna associate lottery', 'committeeId': 4, 'amount': 3000.0}
    ]

    tokens_to_insert = []
    memberships_to_insert = []
    installments_batch = []
    collections_batch = []

    # Phase A: Pass over Scheme Sheets
    for sc in scheme_sheets:
        sh_name = sc['name']
        c_id = sc['committeeId']
        if sh_name not in wb.sheetnames:
            continue
        
        ws = wb[sh_name]
        rows = [r for r in ws.iter_rows(values_only=True) if r and any(x is not None for x in r)]
        if not rows:
            continue

        header_row = rows[0]
        month_cols = []
        for idx in range(6, len(header_row)):
            val = header_row[idx]
            if val is not None:
                d_val = parse_date(val)
                month_cols.append((idx, val, d_val))

        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row or len(row) < 2:
                continue

            raw_token = str(row[0]).strip().split('.')[0] if row[0] is not None else None
            if not raw_token or not raw_token.isdigit():
                continue

            raw_name = clean_name(row[1]) or clean_name(row[2])
            if not raw_name:
                continue

            raw_mobile = clean_phone(row[3]) or clean_phone(row[4])
            raw_ref_name = clean_name(row[2])
            raw_address = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

            key_mob = raw_mobile or f"999{int(raw_token):07d}"
            key_name = raw_name.lower()

            cust_id = existing_cust.get(key_mob) or existing_cust.get(key_name)
            if not cust_id:
                ref_num = f"BS-{c_id}-{raw_token}"
                if ref_num not in new_customers_dict:
                    new_customers_dict[ref_num] = (ref_num, raw_name, key_mob, raw_ref_name, raw_address, key_name)

    # Phase B: Pass over Collection Sheets
    collection_sheets = [
        {'name': 'Daily collection', 'collector': 'Daily Collector'},
        {'name': 'Manager collection', 'collector': 'Manager'},
        {'name': 'Aayush collection', 'collector': 'Aayush'},
        {'name': 'online collection(nikku ji)', 'collector': 'Nikku Ji'},
        {'name': 'recovery collection', 'collector': 'Recovery'}
    ]

    for c_sh in collection_sheets:
        sh_name = c_sh['name']
        if sh_name not in wb.sheetnames:
            continue

        ws = wb[sh_name]
        rows = [r for r in ws.iter_rows(values_only=True) if r and any(x is not None for x in r)]
        if not rows:
            continue

        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row or not row[0]:
                continue

            raw_name = clean_name(row[0])
            if not raw_name:
                continue

            key_name = raw_name.lower()
            cust_id = existing_cust.get(key_name)
            if not cust_id:
                first_word = key_name.split()[0]
                matched_cid = None
                for kn, cid in existing_cust.items():
                    if kn.startswith(first_word):
                        matched_cid = cid
                        break
                if not matched_cid:
                    ref_num = f"CUST-COLL-{c_sh['collector'][:3].upper()}-{r_idx}"
                    if ref_num not in new_customers_dict:
                        new_customers_dict[ref_num] = (ref_num, raw_name, f"9990{r_idx:06d}", None, "", key_name)

    # Bulk Insert All New Customers
    if new_customers_dict:
        print(f"Bulk inserting {len(new_customers_dict)} new customers...")
        cust_values = [
            (ref, name, mob, ref_n, addr, 1, 'active', datetime.now(), datetime.now())
            for ref, (ref, name, mob, ref_n, addr, kn) in new_customers_dict.items()
        ]
        inserted = psycopg2.extras.execute_values(
            cur,
            """
            INSERT INTO customers (reference_number, name, mobile, reference_name, address, branch_id, status, created_at, updated_at)
            VALUES %s
            RETURNING id, reference_number, name, mobile
            """,
            cust_values,
            page_size=1000,
            fetch=True
        )
        for cid, ref, name, mob in inserted:
            if mob: existing_cust[mob] = cid
            if ref: existing_cust[ref] = cid
            if name: existing_cust[name.lower().strip()] = cid
        conn.commit()
        print(f"✓ Customers ready! Total in memory: {len(existing_cust)}")

    # Phase C: Build Tokens, Memberships & Installments
    for sc in scheme_sheets:
        sh_name = sc['name']
        c_id = sc['committeeId']
        if sh_name not in wb.sheetnames:
            continue
        
        ws = wb[sh_name]
        rows = [r for r in ws.iter_rows(values_only=True) if r and any(x is not None for x in r)]
        if not rows:
            continue

        header_row = rows[0]
        month_cols = []
        for idx in range(6, len(header_row)):
            val = header_row[idx]
            if val is not None:
                d_val = parse_date(val)
                month_cols.append((idx, val, d_val))

        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row or len(row) < 2:
                continue

            raw_token = str(row[0]).strip().split('.')[0] if row[0] is not None else None
            if not raw_token or not raw_token.isdigit():
                continue

            raw_name = clean_name(row[1]) or clean_name(row[2])
            if not raw_name:
                continue

            raw_mobile = clean_phone(row[3]) or clean_phone(row[4])
            key_mob = raw_mobile or f"999{int(raw_token):07d}"
            key_name = raw_name.lower()

            cust_id = existing_cust.get(key_mob) or existing_cust.get(key_name)
            if not cust_id:
                continue

            token_key = (raw_token, c_id)
            if token_key in existing_tokens:
                token_id = existing_tokens[token_key]
            else:
                tokens_to_insert.append((raw_token, cust_id, c_id, 'active', datetime.now(), datetime.now()))
                token_id = len(existing_tokens) + len(tokens_to_insert)

            mem_key = (c_id, cust_id, raw_token)
            if mem_key not in existing_memberships:
                memberships_to_insert.append((c_id, cust_id, raw_token, 'active', datetime.now()))
                existing_memberships.add(mem_key)

            for col_idx, raw_col_header, m_date in month_cols:
                if col_idx < len(row):
                    cell_val = str(row[col_idx]).strip() if row[col_idx] is not None else ""
                    if cell_val and cell_val.lower() not in ("none", "-", ""):
                        clean_num = re.sub(r"[^\d.]", "", cell_val)
                        try:
                            amount = float(clean_num) if clean_num and clean_num != "." else sc["amount"]
                        except Exception:
                            amount = sc["amount"]

                        if amount > 100000 or amount <= 0:
                            amount = sc["amount"]

                        if amount > 0:
                            m_num = m_date.month
                            m_yr = m_date.year
                            remarks = f"Installment {raw_col_header}"
                            installments_batch.append((cust_id, token_id, c_id, m_num, m_yr, amount, m_date, 'cash', remarks, datetime.now()))

    # Phase D: Build Collections Batch
    for c_sh in collection_sheets:
        sh_name = c_sh['name']
        if sh_name not in wb.sheetnames:
            continue

        ws = wb[sh_name]
        rows = [r for r in ws.iter_rows(values_only=True) if r and any(x is not None for x in r)]
        if not rows:
            continue

        for r_idx in range(1, len(rows)):
            row = rows[r_idx]
            if not row or not row[0]:
                continue

            raw_name = clean_name(row[0])
            if not raw_name:
                continue

            date_val = row[1] if len(row) > 1 else None
            p_date = parse_date(date_val)

            cash_amt = 0.0
            online_amt = 0.0
            if len(row) > 2 and row[2] is not None:
                try: cash_amt = float(re.sub(r"[^\d.]", "", str(row[2])))
                except Exception: pass
            if len(row) > 3 and row[3] is not None:
                try: online_amt = float(re.sub(r"[^\d.]", "", str(row[3])))
                except Exception: pass

            total_amt = cash_amt + online_amt
            if total_amt <= 0 or total_amt > 1000000:
                continue

            payment_mode = 'upi' if online_amt > 0 else 'cash'
            key_name = raw_name.lower()

            cust_id = existing_cust.get(key_name)
            if not cust_id:
                first_word = key_name.split()[0]
                for kn, cid in existing_cust.items():
                    if kn.startswith(first_word):
                        cust_id = cid
                        break
            if not cust_id:
                ref_num = f"CUST-COLL-{c_sh['collector'][:3].upper()}-{r_idx}"
                cust_id = existing_cust.get(ref_num)

            if cust_id:
                receipt_no = f"REC-{c_sh['collector'][:3].upper()}-{r_idx}"
                notes = f"Bissi collection via {c_sh['collector']} (Receipt: {receipt_no})"
                collections_batch.append((cust_id, 1, total_amt, payment_mode, receipt_no, notes, 'verified', p_date, datetime.now()))

    # Execute Bulk Inserts
    if tokens_to_insert:
        print(f"Bulk inserting {len(tokens_to_insert)} tokens...")
        psycopg2.extras.execute_values(cur, "INSERT INTO tokens (token_number, customer_id, committee_id, status, created_at, updated_at) VALUES %s", tokens_to_insert, page_size=1000)

    if memberships_to_insert:
        print(f"Bulk inserting {len(memberships_to_insert)} memberships...")
        psycopg2.extras.execute_values(cur, "INSERT INTO committee_members (committee_id, customer_id, token_number, status, joined_at) VALUES %s", memberships_to_insert, page_size=1000)

    if installments_batch:
        print(f"Bulk inserting {len(installments_batch)} installments...")
        psycopg2.extras.execute_values(cur, "INSERT INTO installments (customer_id, token_id, committee_id, month, year, amount, payment_date, payment_mode, remarks, created_at) VALUES %s", installments_batch, page_size=1000)

    if collections_batch:
        print(f"Bulk inserting {len(collections_batch)} collections...")
        psycopg2.extras.execute_values(cur, "INSERT INTO collections (customer_id, branch_id, amount, payment_mode, receipt_number, notes, verification_status, collected_at, created_at) VALUES %s", collections_batch, page_size=1000)

    conn.commit()
    cur.close()
    conn.close()
    print(f"🎉 MASTER BULK IMPORT FINISHED! ({len(installments_batch)} Installments, {len(collections_batch)} Collections inserted).")

if __name__ == "__main__":
    run_master_import()

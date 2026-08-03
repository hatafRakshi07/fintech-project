"""
Import missing JSK/blank tokens:
- Name = reference person if available, else "Unknown (Token #N)"
- Import all their payment collections
- They'll show in pending reports automatically
"""
import openpyxl
import psycopg2
import psycopg2.extras
import re
from datetime import date, datetime

WORKBOOK_PATH = 'C:/Users/iSN_kota_T52/Downloads/Bissi folder (1).xlsx'
DB_URL = 'postgresql://neondb_owner:npg_qSQN29ZxTKzt@ep-frosty-cloud-at51tjed.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require'

ORG_ID = '00000000-0000-0000-0000-000000000001'

SCHEMES = [
    ('Hare ka sahara bissi 20 date',    'BISSI-1', '11111111-1111-1111-1111-111111111111', 7,  2500),
    ('Shree Krishna associate lottery', 'BISSI-2', '22222222-2222-2222-2222-222222222222', 6,  3000),
    ('Pyare mohan 15 date',             'BISSI-3', '33333333-3333-3333-3333-333333333333', 9,  3000),
    ('Sawariya seth 5 date',            'BISSI-4', 'a3d68b9c-63df-4884-a5ad-eb8a17e3be31', 7,  3000),
]

def clean_ref(ref):
    if not ref or str(ref).strip() in ['-', '', 'None']:
        return None
    r = str(ref).strip()
    if r.lower() in ['jsk', 'jsk dwari bai', 'tilumal jsk', 'jsk (harish ji)', 'jsk dwari bai']:
        return None
    return r[:60]

def make_customer_name(token_num, ref, scheme_code):
    r = clean_ref(ref)
    if r:
        return f"{r} (Token #{token_num})"
    return f"Unknown Member (Token #{token_num})"

def main():
    print('Loading workbook...')
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    conn = psycopg2.connect(DB_URL)
    cur  = conn.cursor()

    total_customers = 0
    total_tokens    = 0
    total_payments  = 0

    for sheet_name, code, comm_uuid, data_col, installment in SCHEMES:
        ws = wb[sheet_name]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        date_cols = [(i, v) for i, v in enumerate(header) if v and hasattr(v, 'year')]

        # Fetch already-imported tokens for this committee
        cur.execute('SELECT normalized_token_number FROM tokens WHERE committee_id = %s', [comm_uuid])
        existing = {row[0] for row in cur.fetchall()}

        scheme_cust = scheme_tok = scheme_pay = 0

        for r in range(2, ws.max_row + 1):
            tok_val = ws.cell(r, 1).value
            if not tok_val:
                continue
            try:
                tok_num = int(float(str(tok_val)))
            except:
                continue

            if tok_num in existing:
                continue  # already imported

            name_val = str(ws.cell(r, 2).value or '').strip()
            ref_val  = str(ws.cell(r, 3).value or '').strip()

            # Collect payment data
            payments = []
            for col_idx, col_date in date_cols:
                v = ws.cell(r, col_idx + 1).value
                if v and isinstance(v, (int, float)) and v > 0:
                    payments.append((col_date.date() if isinstance(col_date, datetime) else col_date, float(v)))

            # Skip completely blank rows with no payments
            if not payments:
                continue

            # Build customer name
            cust_name = make_customer_name(tok_num, ref_val or name_val, code)

            # 1. Create customer
            cur.execute("""
                INSERT INTO customers (organization_id, name, mobile, status, created_at, updated_at)
                VALUES (%s, %s, '0000000000', 'ACTIVE', NOW(), NOW())
                RETURNING id
            """, [ORG_ID, cust_name])
            cust_uuid = cur.fetchone()[0]
            scheme_cust += 1

            # 2. Create token (display_token is a generated column, skip it)
            cur.execute("""
                INSERT INTO tokens
                    (organization_id, committee_id, customer_id,
                     raw_token_number, normalized_token_number, duplicate_suffix,
                     status, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, '', 'ACTIVE', NOW(), NOW())
                RETURNING id
            """, [ORG_ID, comm_uuid, cust_uuid, str(tok_num), tok_num])
            tok_uuid = cur.fetchone()[0]
            scheme_tok += 1

            # 3. Insert collections
            if payments:
                psycopg2.extras.execute_values(cur, """
                    INSERT INTO collections
                        (customer_id, committee_id, amount, payment_mode,
                         collected_at, created_at, verification_status,
                         customer_uuid, committee_uuid, token_uuid, notes)
                    VALUES %s
                """, [
                    (1, 1, amt, 'cash',
                     pay_date, datetime.now(), 'verified',
                     str(cust_uuid), comm_uuid, str(tok_uuid),
                     f'Token {tok_num} (JSK/Unknown)')
                    for pay_date, amt in payments
                ], template="(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")
                scheme_pay += len(payments)

        conn.commit()
        print(f'{code}: +{scheme_cust} customers, +{scheme_tok} tokens, +{scheme_pay} payments')
        total_customers += scheme_cust
        total_tokens    += scheme_tok
        total_payments  += scheme_pay

    print(f'\nDONE: {total_customers} customers, {total_tokens} tokens, {total_payments} payments added')

    # Verify
    cur.execute('SELECT count(*) FROM collections WHERE committee_uuid IS NOT NULL')
    print(f'Total collections now: {cur.fetchone()[0]}')
    cur.execute('SELECT count(*) FROM tokens')
    print(f'Total tokens now: {cur.fetchone()[0]}')

    conn.close()
    wb.close()

if __name__ == '__main__':
    main()
